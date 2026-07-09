from __future__ import annotations

import argparse
import csv
import json
import zipfile
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "euc-kr")


def read_csv_rows(path: Path) -> tuple[list[dict[str, str]], str]:
    errors: list[str] = []
    for encoding in CSV_ENCODINGS:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                rows = [{(key or "").strip(): (value or "").strip() for key, value in row.items()} for row in reader]
            return rows, encoding
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise RuntimeError(f"Unable to decode {path}: {' | '.join(errors)}")


def summarize_values(rows: list[dict[str, str]], columns: list[str]) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for column in columns:
        values = [row.get(column, "") for row in rows]
        non_empty = [value for value in values if value]
        counter = Counter(non_empty)
        summary[column] = {
            "non_empty": len(non_empty),
            "unique": len(counter),
            "top_values": [{"value": value, "count": count} for value, count in counter.most_common(5)],
        }
    return summary


def inspect_csv(path: Path) -> dict[str, Any]:
    rows, encoding = read_csv_rows(path)
    columns = list(rows[0].keys()) if rows else []
    return {
        "type": "csv",
        "path": str(path),
        "encoding": encoding,
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "sample_rows": rows[:3],
        "value_summary": summarize_values(rows, columns[: min(len(columns), 8)]),
    }


def inspect_zip(path: Path) -> dict[str, Any]:
    members: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            member: dict[str, Any] = {
                "name": info.filename,
                "size": info.file_size,
            }
            lower_name = info.filename.lower()
            if lower_name.endswith(".csv"):
                with archive.open(info, "r") as raw_handle:
                    raw_bytes = raw_handle.read()
                parsed = None
                parse_errors: list[str] = []
                for encoding in CSV_ENCODINGS:
                    try:
                        text = raw_bytes.decode(encoding)
                        reader = csv.DictReader(text.splitlines())
                        rows = [{(key or "").strip(): (value or "").strip() for key, value in row.items()} for row in reader]
                        parsed = {
                            "encoding": encoding,
                            "row_count": len(rows),
                            "column_count": len(rows[0].keys()) if rows else 0,
                            "columns": list(rows[0].keys()) if rows else [],
                            "sample_rows": rows[:2],
                        }
                        break
                    except UnicodeDecodeError as exc:
                        parse_errors.append(f"{encoding}: {exc}")
                if parsed is not None:
                    member["csv_profile"] = parsed
                else:
                    member["csv_profile_error"] = " | ".join(parse_errors)
            elif lower_name.endswith(".xls"):
                with archive.open(info, "r") as raw_handle:
                    workbook = xlrd.open_workbook(file_contents=raw_handle.read())
                sheets = [
                    {
                        "name": sheet.name,
                        "row_count": sheet.nrows,
                        "column_count": sheet.ncols,
                        "header": [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)] if sheet.nrows else [],
                    }
                    for sheet in workbook.sheets()
                ]
                non_empty = [sheet for sheet in sheets if sheet["row_count"] or sheet["column_count"]]
                member["xls_profile"] = {
                    "sheet_count": workbook.nsheets,
                    "sheets": sheets[:10],
                    "non_empty_sheets": non_empty[:10],
                }
            elif lower_name.endswith(".xlsx"):
                with archive.open(info, "r") as raw_handle:
                    workbook = load_workbook(filename=BytesIO(raw_handle.read()), read_only=True, data_only=True)
                sheets = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    header = []
                    if sheet.max_row and sheet.max_column:
                        header = [
                            "" if value is None else str(value).strip()
                            for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                        ]
                    sheets.append(
                        {
                            "name": sheet_name,
                            "row_count": sheet.max_row,
                            "column_count": sheet.max_column,
                            "header": header,
                        }
                    )
                non_empty = [sheet for sheet in sheets if sheet["row_count"] or sheet["column_count"]]
                member["xlsx_profile"] = {
                    "sheet_count": len(workbook.sheetnames),
                    "sheets": sheets[:10],
                    "non_empty_sheets": non_empty[:10],
                }
            members.append(member)
    return {
        "type": "zip",
        "path": str(path),
        "member_count": len(members),
        "members": members,
    }


def inspect_path(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return inspect_csv(path)
    if suffix == ".zip":
        return inspect_zip(path)
    return {
        "type": "unknown",
        "path": str(path),
        "suffix": suffix,
        "size": path.stat().st_size,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect public data files for migration value audit.")
    parser.add_argument("paths", nargs="+", help="CSV or ZIP file paths to inspect")
    parser.add_argument("--output", required=True, help="Output JSON path")
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "files": [inspect_path(Path(raw_path)) for raw_path in args.paths],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
