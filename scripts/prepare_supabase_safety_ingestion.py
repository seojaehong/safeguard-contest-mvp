from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import time
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {".csv", ".hwp", ".hwpx", ".pdf", ".pptx", ".xls", ".xlsx", ".zip", ".jpg", ".jpeg", ".png"}
DOWNLOADS_SINCE_DEFAULT = "2026-05-03T18:45:00"
SOURCE_BATCH = "supabase-safety-final-candidates-2026-05-03"

SAFETY_KEYWORDS = [
    "TBM",
    "Tool Box",
    "위험성평가",
    "작업계획",
    "작업계획서",
    "허가",
    "허가서",
    "안전교육",
    "교육일지",
    "사진",
    "증빙",
    "비상",
    "대응",
    "기술지원",
    "안전보건",
    "산업안전",
    "점검표",
    "자율점검",
    "일지",
    "서식",
    "중대재해",
    "추락",
    "끼임",
    "충돌",
    "전도",
    "화재",
    "폭발",
    "질식",
    "붕괴",
    "한랭",
    "폭염",
    "장마",
    "해빙기",
    "기계",
    "장비",
    "비계",
]

DOCUMENT_KIND_LABELS = {
    "tbm": "TBM",
    "risk_assessment": "위험성평가",
    "work_plan": "작업계획서",
    "permit": "허가서",
    "safety_education_log": "안전교육일지",
    "photo_evidence": "사진증빙",
    "emergency_response": "비상대응",
    "technical_reference": "내부 지식DB",
}

FIELD_PATTERNS = [
    "사업장명",
    "공사명",
    "작업명",
    "작업일",
    "작업장소",
    "작업내용",
    "위험요인",
    "감소대책",
    "안전대책",
    "점검사항",
    "확인",
    "서명",
    "작성자",
    "관리감독자",
    "교육대상",
    "교육내용",
    "참석자",
    "허가",
    "첨부",
    "사진",
    "조치",
]


@dataclass(frozen=True)
class CandidateFile:
    source_root: str
    path: Path
    extension: str
    size_bytes: int
    modified_at: str
    sha256: str
    normalized_title: str
    duplicate_of_existing: bool
    duplicate_group: str | None
    duplicate_reason: str | None


@dataclass(frozen=True)
class CandidateRecord:
    id: str
    source_id: str
    source_file: str
    source_type: str
    form_type: str
    document_kind: str
    page_or_sheet: str
    extracted_text: str
    section_title: str
    field_labels: list[str]
    recommended_use: str
    confidence: float
    needs_manual_review: bool
    parser: str
    failure_reason: str | None
    duplicate_group: str | None
    payload: dict[str, Any]


@dataclass(frozen=True)
class SourcePayload:
    id: str
    source_group: str
    source_type: str
    agency: str
    title: str
    source_path: str
    origin_url: str | None
    file_format: str
    published_at: str | None
    metadata: dict[str, Any]


@dataclass(frozen=True)
class ItemPayload:
    id: str
    source_id: str
    item_type: str
    category: str | None
    subcategory: str | None
    title: str
    summary: str
    body: str
    keywords: list[str]
    risk_tags: list[str]
    primary_documents: list[str]
    controls: list[str]
    payload: dict[str, Any]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def compact_text(value: str, limit: int = 1800) -> str:
    return re.sub(r"\s+", " ", value or "").strip()[:limit].rstrip()


def slugify(value: str, max_length: int = 100) -> str:
    normalized = re.sub(r"[^0-9A-Za-z가-힣]+", "-", value.strip()).strip("-").lower()
    if not normalized:
        return hashlib.sha1(value.encode("utf-8", errors="ignore")).hexdigest()[:12]
    if len(normalized) <= max_length:
        return normalized
    digest = hashlib.sha1(value.encode("utf-8", errors="ignore")).hexdigest()[:10]
    return f"{normalized[:max_length - 11].rstrip('-')}-{digest}"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_title(path: Path) -> str:
    title = path.stem.lower()
    title = re.sub(r"\(\d+\)$", "", title).strip()
    title = title.replace("_", " ")
    return re.sub(r"\s+", " ", title)


def is_relevant(path: Path) -> bool:
    text = str(path)
    return any(keyword.lower() in text.lower() for keyword in SAFETY_KEYWORDS)


def load_existing_bundle(path: Path) -> tuple[set[str], set[str]]:
    if not path.exists():
        return set(), set()
    data = json.loads(path.read_text(encoding="utf-8"))
    inventory = data.get("inventory", [])
    hashes = {str(item.get("sha256")) for item in inventory if item.get("sha256")}
    titles = {normalize_title(Path(str(item.get("path", item.get("name", ""))))) for item in inventory if item.get("name") or item.get("path")}
    return hashes, titles


def find_files(downloads_dir: Path, onedrive_dir: Path, downloads_since: datetime) -> list[Path]:
    files: list[Path] = []
    for path in downloads_dir.iterdir():
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if path.name.lower().startswith("client_secret"):
            continue
        modified = datetime.fromtimestamp(path.stat().st_mtime)
        if modified >= downloads_since and is_relevant(path):
            files.append(path)
    for path in onedrive_dir.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if is_relevant(path):
            files.append(path)
    return sorted(set(files), key=lambda item: (str(item.parent), item.name))


def read_pdf_records(path: Path, max_pages: int) -> tuple[list[tuple[str, str, str]], str, str | None]:
    try:
        reader = PdfReader(str(path))
        records: list[tuple[str, str, str]] = []
        for index, page in enumerate(reader.pages[:max_pages], start=1):
            text = compact_text(page.extract_text() or "", 2500)
            if text:
                records.append((f"page:{index}", infer_section_title(path.stem, text), text))
        if not records:
            return [], "pypdf", "no extractable text; OCR required"
        return records, "pypdf", None
    except Exception as exc:
        return [], "pypdf", str(exc)


def read_hwpx_records(path: Path) -> tuple[list[tuple[str, str, str]], str, str | None]:
    try:
        records: list[tuple[str, str, str]] = []
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                lower_name = name.lower()
                if not lower_name.endswith(".xml") or "/section" not in lower_name:
                    continue
                root = ElementTree.fromstring(archive.read(name))
                texts = [element.text.strip() for element in root.iter() if element.text and element.text.strip()]
                text = compact_text("\n".join(texts), 2500)
                if text:
                    records.append((name, infer_section_title(path.stem, text), text))
        if not records:
            return [], "zip+xml", "no extractable hwp section text"
        return records, "zip+xml", None
    except Exception as exc:
        return [], "zip+xml", str(exc)


def read_xlsx_records(path: Path, max_rows: int) -> tuple[list[tuple[str, str, str]], str, str | None]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        records: list[tuple[str, str, str]] = []
        for sheet_name in workbook.sheetnames:
            lines: list[str] = []
            sheet = workbook[sheet_name]
            for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if index > max_rows:
                    break
                values = [compact_text(str(cell), 220) for cell in row if cell is not None and compact_text(str(cell), 220)]
                if values:
                    lines.append(" | ".join(values))
            text = compact_text("\n".join(lines), 2500)
            if text:
                records.append((f"sheet:{sheet_name}", sheet_name, text))
        if not records:
            return [], "openpyxl", "no non-empty sheets"
        return records, "openpyxl", None
    except Exception as exc:
        return [], "openpyxl", str(exc)


def read_csv_records(path: Path, max_rows: int) -> tuple[list[tuple[str, str, str]], str, str | None]:
    encodings = ["utf-8-sig", "cp949", "euc-kr"]
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                rows = [" | ".join(row) for _, row in zip(range(max_rows), csv.reader(file))]
            text = compact_text("\n".join(rows), 2500)
            if text:
                return [("rows:1-60", path.stem, text)], f"csv:{encoding}", None
            return [], f"csv:{encoding}", "no csv rows"
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            return [], f"csv:{encoding}", str(exc)
    return [], "csv", "encoding detection failed"


def read_zip_records(path: Path) -> tuple[list[tuple[str, str, str]], str, str | None]:
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
        text = compact_text("\n".join(names), 2500)
        return [("zip:listing", path.stem, text)] if text else [], "zip-listing", None if text else "empty zip"
    except Exception as exc:
        return [], "zip-listing", str(exc)


def read_pptx_records(path: Path) -> tuple[list[tuple[str, str, str]], str, str | None]:
    try:
        records: list[tuple[str, str, str]] = []
        with zipfile.ZipFile(path) as archive:
            slide_names = sorted(name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
            for index, name in enumerate(slide_names[:20], start=1):
                root = ElementTree.fromstring(archive.read(name))
                texts = [element.text.strip() for element in root.iter() if element.text and element.text.strip()]
                text = compact_text("\n".join(texts), 2200)
                if text:
                    records.append((f"slide:{index}", infer_section_title(path.stem, text), text))
        if not records:
            return [], "pptx-zip+xml", "no extractable slide text"
        return records, "pptx-zip+xml", None
    except Exception as exc:
        return [], "pptx-zip+xml", str(exc)


def extract_records(path: Path, max_pdf_pages: int, max_sheet_rows: int) -> tuple[list[tuple[str, str, str]], str, str | None]:
    extension = path.suffix.lower()
    if extension == ".pdf":
        return read_pdf_records(path, max_pdf_pages)
    if extension == ".hwpx":
        return read_hwpx_records(path)
    if extension == ".xlsx":
        return read_xlsx_records(path, max_sheet_rows)
    if extension == ".csv":
        return read_csv_records(path, max_sheet_rows)
    if extension == ".zip":
        return read_zip_records(path)
    if extension == ".pptx":
        return read_pptx_records(path)
    if extension in {".jpg", ".jpeg", ".png"}:
        return [("image:metadata", path.stem, f"Image evidence file: {path.name}")], "image-metadata", None
    if extension in {".hwp", ".xls"}:
        return [], "metadata-only", f"{extension} binary parser not enabled; use HWPX/XLSX/OCR conversion before production load"
    return [], "metadata-only", f"unsupported extension: {extension}"


def infer_document_kind(name: str, text: str) -> str:
    combined = f"{name}\n{text}".lower()
    if "tbm" in combined or "tool box" in combined or "작업 전 안전점검회의" in combined:
        return "tbm"
    if "위험성평가" in combined or "위험요인" in combined or "자율점검표" in combined:
        return "risk_assessment"
    if "작업계획서" in combined or "표준 작업계획서" in combined:
        return "work_plan"
    if "허가" in combined or "이행각서" in combined or "지정서" in combined or "문진표" in combined:
        return "permit"
    if "교육" in combined or "교육일지" in combined:
        return "safety_education_log"
    if "사진" in combined or "증빙" in combined or "image evidence" in combined:
        return "photo_evidence"
    if "비상" in combined or "대응" in combined or "응급" in combined:
        return "emergency_response"
    return "technical_reference"


def infer_form_type(document_kind: str, path: Path) -> str:
    if path.suffix.lower() in {".hwpx", ".hwp", ".xlsx", ".xls"} and document_kind != "technical_reference":
        return "form_template"
    if document_kind in {"tbm", "risk_assessment", "work_plan", "permit", "safety_education_log"}:
        return "field_form_or_guide"
    return "knowledge_reference"


def infer_section_title(default_title: str, text: str) -> str:
    for line in re.split(r"[\r\n]+", text):
        cleaned = compact_text(line, 90)
        if 4 <= len(cleaned) <= 90:
            return cleaned
    return default_title


def extract_field_labels(text: str) -> list[str]:
    labels: list[str] = []
    for label in FIELD_PATTERNS:
        if label in text:
            labels.append(label)
    colon_labels = re.findall(r"([가-힣A-Za-z0-9·/\s]{2,18})\s*[:：]", text)
    for label in colon_labels:
        cleaned = compact_text(label, 24)
        if cleaned and cleaned not in labels:
            labels.append(cleaned)
    return labels[:18]


def recommended_use(document_kind: str, form_type: str) -> str:
    if document_kind in {"tbm", "risk_assessment", "work_plan", "permit", "safety_education_log", "photo_evidence", "emergency_response"}:
        return "submission_output" if form_type == "form_template" or document_kind in {"tbm", "risk_assessment", "work_plan", "permit"} else "internal_knowledge_db"
    return "internal_knowledge_db"


def confidence_score(text: str, parser: str, document_kind: str, duplicate: bool) -> float:
    score = 0.35
    if text:
        score += 0.25
    if len(text) > 300:
        score += 0.15
    if parser not in {"metadata-only", "image-metadata"}:
        score += 0.15
    if document_kind != "technical_reference":
        score += 0.08
    if duplicate:
        score -= 0.08
    return round(max(0.05, min(score, 0.95)), 2)


def keywords_from_text(name: str, text: str) -> list[str]:
    combined = f"{name}\n{text}"
    keywords = [keyword for keyword in SAFETY_KEYWORDS if keyword.lower() in combined.lower()]
    return sorted(set(keywords), key=keywords.index)[:24]


def primary_documents(document_kind: str, use: str) -> list[str]:
    mapping = {
        "tbm": ["TBM 브리핑", "TBM 기록"],
        "risk_assessment": ["위험성평가표"],
        "work_plan": ["작업계획서"],
        "permit": ["허가서/첨부"],
        "safety_education_log": ["안전보건교육 기록"],
        "photo_evidence": ["사진/증빙"],
        "emergency_response": ["비상대응 절차"],
        "technical_reference": ["위험성평가표", "TBM 브리핑", "안전보건교육 기록"],
    }
    docs = mapping.get(document_kind, mapping["technical_reference"])
    if use == "internal_knowledge_db" and document_kind != "technical_reference":
        return docs + ["내부 지식DB"]
    return docs


def controls_for(document_kind: str, keywords: list[str]) -> list[str]:
    controls: list[str] = []
    if document_kind == "tbm":
        controls.append("작업 전 TBM 참석자 서명과 위험요인 복창 확인")
    if document_kind == "risk_assessment":
        controls.append("유해·위험요인별 감소대책과 잔여위험 확인")
    if document_kind == "work_plan":
        controls.append("작업순서·장비·인원·작업중지 기준 확인")
    if document_kind == "permit":
        controls.append("허가대상·첨부서류·종료확인 서명란 확인")
    if document_kind == "safety_education_log":
        controls.append("교육대상·교육내용·이수 서명 확인")
    if document_kind == "photo_evidence":
        controls.append("개선 전후 사진과 촬영시각 증빙 확보")
    if document_kind == "emergency_response":
        controls.append("비상연락망·대피·응급조치 절차 확인")
    if any(keyword in keywords for keyword in ["추락", "비계"]):
        controls.append("작업발판·난간·안전대 체결 확인")
    if any(keyword in keywords for keyword in ["기계", "장비", "충돌", "끼임"]):
        controls.append("장비 동선 분리와 신호수 배치 확인")
    if any(keyword in keywords for keyword in ["폭염", "한랭", "장마", "해빙기"]):
        controls.append("기상특보와 작업중지 기준 공유")
    if not controls:
        controls.append("작업 전 유해·위험요인 확인")
    return sorted(set(controls), key=controls.index)


def source_root_label(path: Path, downloads_dir: Path, onedrive_dir: Path) -> str:
    try:
        path.relative_to(downloads_dir)
        return "downloads"
    except ValueError:
        pass
    try:
        path.relative_to(onedrive_dir)
        return "onedrive-industrial-safety"
    except ValueError:
        return "external"


def build_candidates(paths: Iterable[Path], downloads_dir: Path, onedrive_dir: Path, existing_hashes: set[str], existing_titles: set[str]) -> list[CandidateFile]:
    path_list = list(paths)
    hashes = {path: sha256_file(path) for path in path_list}
    hash_groups: dict[str, list[Path]] = {}
    title_groups: dict[str, list[Path]] = {}
    for path, digest in hashes.items():
        hash_groups.setdefault(digest, []).append(path)
        title_groups.setdefault(normalize_title(path), []).append(path)
    candidates: list[CandidateFile] = []
    for path in path_list:
        stat = path.stat()
        digest = hashes[path]
        title = normalize_title(path)
        duplicate_group: str | None = None
        duplicate_reason: str | None = None
        if digest in existing_hashes:
            duplicate_group = f"existing-sha256:{digest[:12]}"
            duplicate_reason = "same hash as existing download-safety-bundle"
        elif title in existing_titles:
            duplicate_group = f"existing-title:{slugify(title, 70)}"
            duplicate_reason = "same normalized title as existing download-safety-bundle"
        elif len(hash_groups[digest]) > 1:
            duplicate_group = f"sha256:{digest[:12]}"
            duplicate_reason = "same hash within current scan"
        elif len(title_groups[title]) > 1:
            duplicate_group = f"same-title:{slugify(title, 70)}"
            duplicate_reason = "same normalized title within current scan"
        candidates.append(CandidateFile(
            source_root=source_root_label(path, downloads_dir, onedrive_dir),
            path=path,
            extension=path.suffix.lower(),
            size_bytes=stat.st_size,
            modified_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
            sha256=digest,
            normalized_title=title,
            duplicate_of_existing=digest in existing_hashes or title in existing_titles,
            duplicate_group=duplicate_group,
            duplicate_reason=duplicate_reason,
        ))
    return candidates


def build_records(candidate: CandidateFile, max_pdf_pages: int, max_sheet_rows: int) -> tuple[SourcePayload, list[CandidateRecord], list[ItemPayload], bool]:
    source_id = f"{candidate.source_root}-{slugify(candidate.path.stem)}-{candidate.sha256[:10]}"
    tuples, parser, failure_reason = extract_records(candidate.path, max_pdf_pages, max_sheet_rows)
    if not tuples:
        tuples = [("file:metadata", candidate.path.stem, "")]
    records: list[CandidateRecord] = []
    items: list[ItemPayload] = []
    success = failure_reason is None and any(text for _, _, text in tuples)
    for index, (page_or_sheet, section_title, text) in enumerate(tuples, start=1):
        document_kind = infer_document_kind(candidate.path.name, text)
        form_type = infer_form_type(document_kind, candidate.path)
        use = recommended_use(document_kind, form_type)
        labels = extract_field_labels(text)
        keywords = keywords_from_text(candidate.path.name, text)
        confidence = confidence_score(text, parser, document_kind, bool(candidate.duplicate_group))
        needs_review = bool(failure_reason) or confidence < 0.72 or candidate.duplicate_of_existing or not labels
        record = CandidateRecord(
            id=f"{source_id}-record-{index:04d}",
            source_id=source_id,
            source_file=str(candidate.path),
            source_type=candidate.extension.lstrip("."),
            form_type=form_type,
            document_kind=document_kind,
            page_or_sheet=page_or_sheet,
            extracted_text=text,
            section_title=section_title,
            field_labels=labels,
            recommended_use=use,
            confidence=confidence,
            needs_manual_review=needs_review,
            parser=parser,
            failure_reason=failure_reason,
            duplicate_group=candidate.duplicate_group,
            payload={
                "sourceRoot": candidate.source_root,
                "modifiedAt": candidate.modified_at,
                "sha256": candidate.sha256,
                "duplicateReason": candidate.duplicate_reason,
            },
        )
        records.append(record)
        if text:
            items.append(ItemPayload(
                id=f"{record.id}-item",
                source_id=source_id,
                item_type=document_kind,
                category=DOCUMENT_KIND_LABELS.get(document_kind, "내부 지식DB"),
                subcategory=form_type,
                title=f"{candidate.path.stem} - {section_title}",
                summary=compact_text(text, 320),
                body=text,
                keywords=keywords,
                risk_tags=[keyword for keyword in keywords if keyword not in {"TBM", "Tool Box", "위험성평가", "작업계획", "안전교육"}],
                primary_documents=primary_documents(document_kind, use),
                controls=controls_for(document_kind, keywords),
                payload={key: value for key, value in asdict(record).items() if key not in {"id", "source_id"}},
            ))
    source = SourcePayload(
        id=source_id,
        source_group=SOURCE_BATCH,
        source_type=records[0].document_kind if records else "technical_reference",
        agency="uploaded/consulting safety source",
        title=candidate.path.stem,
        source_path=str(candidate.path),
        origin_url=None,
        file_format=candidate.extension.lstrip("."),
        published_at=None,
        metadata={
            "sourceRoot": candidate.source_root,
            "sizeBytes": candidate.size_bytes,
            "modifiedAt": candidate.modified_at,
            "sha256": candidate.sha256,
            "normalizedTitle": candidate.normalized_title,
            "parser": parser,
            "parseSuccess": success,
            "failureReason": failure_reason,
            "duplicateGroup": candidate.duplicate_group,
            "duplicateReason": candidate.duplicate_reason,
            "duplicateOfExistingBundle": candidate.duplicate_of_existing,
            "recordCount": len(records),
        },
    )
    return source, records, items, success


def sql_literal(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        return "array[" + ",".join(sql_literal(item) for item in value) + "]::text[]"
    if isinstance(value, dict):
        return "'" + json.dumps(value, ensure_ascii=False).replace("'", "''") + "'::jsonb"
    return "'" + str(value).replace("'", "''") + "'"


def build_upsert_sql(sources: list[SourcePayload], items: list[ItemPayload], run_record: dict[str, Any]) -> str:
    source_values = ",\n".join(
        "(" + ",".join([
            sql_literal(source.id),
            sql_literal(source.source_group),
            sql_literal(source.source_type),
            sql_literal(source.agency),
            sql_literal(source.title),
            sql_literal(source.source_path),
            sql_literal(source.origin_url),
            sql_literal(source.file_format),
            sql_literal(source.published_at),
            sql_literal(source.metadata),
        ]) + ")"
        for source in sources
    )
    item_values = ",\n".join(
        "(" + ",".join([
            sql_literal(item.id),
            sql_literal(item.source_id),
            sql_literal(item.item_type),
            sql_literal(item.category),
            sql_literal(item.subcategory),
            sql_literal(item.title),
            sql_literal(item.summary),
            sql_literal(item.body),
            sql_literal(item.keywords),
            sql_literal(item.risk_tags),
            sql_literal(item.primary_documents),
            sql_literal(item.controls),
            sql_literal(item.payload),
        ]) + ")"
        for item in items
    )
    parts = [
        "-- REVIEW ONLY: generated upsert SQL. This script was not executed.",
        "insert into safety_reference_sources (id, source_group, source_type, agency, title, source_path, origin_url, file_format, published_at, metadata)",
        f"values\n{source_values}",
        "on conflict (id) do update set",
        "  source_group = excluded.source_group,",
        "  source_type = excluded.source_type,",
        "  agency = excluded.agency,",
        "  title = excluded.title,",
        "  source_path = excluded.source_path,",
        "  origin_url = excluded.origin_url,",
        "  file_format = excluded.file_format,",
        "  published_at = excluded.published_at,",
        "  metadata = excluded.metadata,",
        "  updated_at = now();",
    ]
    if item_values:
        parts.extend([
            "",
            "insert into safety_reference_items (id, source_id, item_type, category, subcategory, title, summary, body, keywords, risk_tags, primary_documents, controls, payload)",
            f"values\n{item_values}",
            "on conflict (id) do update set",
            "  source_id = excluded.source_id,",
            "  item_type = excluded.item_type,",
            "  category = excluded.category,",
            "  subcategory = excluded.subcategory,",
            "  title = excluded.title,",
            "  summary = excluded.summary,",
            "  body = excluded.body,",
            "  keywords = excluded.keywords,",
            "  risk_tags = excluded.risk_tags,",
            "  primary_documents = excluded.primary_documents,",
            "  controls = excluded.controls,",
            "  payload = excluded.payload,",
            "  updated_at = now();",
        ])
    parts.extend([
        "",
        "insert into safety_reference_ingestion_runs (source_batch, source_count, item_count, success_count, failure_count, elapsed_ms, report_path, status, details)",
        "values (" + ",".join([
            sql_literal(run_record["source_batch"]),
            str(run_record["source_count"]),
            str(run_record["item_count"]),
            str(run_record["success_count"]),
            str(run_record["failure_count"]),
            str(run_record["elapsed_ms"]),
            sql_literal(run_record["report_path"]),
            sql_literal(run_record["status"]),
            sql_literal(run_record["details"]),
        ]) + ");",
    ])
    return "\n".join(parts) + "\n"


def build_migration_candidate_sql() -> str:
    return """-- MIGRATION CANDIDATE ONLY. Do not apply without review.
-- Existing tables safety_reference_sources/items/ingestion_runs can carry this bundle through item.payload.
-- This optional table is recommended only if form-level QA needs first-class columns.

create table if not exists safety_reference_form_records (
  id text primary key,
  source_id text not null references safety_reference_sources(id) on delete cascade,
  source_file text not null,
  source_type text not null,
  form_type text not null,
  document_kind text not null check (document_kind in (
    'tbm',
    'risk_assessment',
    'work_plan',
    'permit',
    'safety_education_log',
    'photo_evidence',
    'emergency_response',
    'technical_reference'
  )),
  page_or_sheet text not null,
  extracted_text text not null default '',
  section_title text not null default '',
  field_labels text[] not null default '{}'::text[],
  recommended_use text not null check (recommended_use in ('submission_output', 'internal_knowledge_db')),
  confidence numeric(4,2) not null default 0,
  needs_manual_review boolean not null default true,
  parser text not null,
  failure_reason text,
  duplicate_group text,
  payload jsonb not null default '{}'::jsonb,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);

create index if not exists idx_safety_reference_form_records_source_id on safety_reference_form_records(source_id);
create index if not exists idx_safety_reference_form_records_document_kind on safety_reference_form_records(document_kind);
create index if not exists idx_safety_reference_form_records_recommended_use on safety_reference_form_records(recommended_use);
create index if not exists idx_safety_reference_form_records_field_labels on safety_reference_form_records using gin(field_labels);
create index if not exists idx_safety_reference_form_records_search on safety_reference_form_records using gin(
  to_tsvector('simple', coalesce(section_title, '') || ' ' || coalesce(extracted_text, ''))
);

alter table safety_reference_form_records enable row level security;

create policy "public can read safety reference form records"
  on safety_reference_form_records for select
  using (true);
"""


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def count_by(values: Iterable[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for value in values:
        counts[value] = counts.get(value, 0) + 1
    return dict(sorted(counts.items()))


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare final Supabase dry-run payloads for safety forms and references.")
    parser.add_argument("--downloads-dir", default=str(Path.home() / "Downloads"))
    parser.add_argument("--onedrive-dir", default=r"C:\Users\iceam\OneDrive\_30_컨설팅\2025\산업안전")
    parser.add_argument("--existing-bundle", default="evaluation/download-safety-bundle-2026-05-03/parse-report.json")
    parser.add_argument("--out-dir", default="evaluation/supabase-safety-ingestion-ready-2026-05-03")
    parser.add_argument("--downloads-since", default=DOWNLOADS_SINCE_DEFAULT)
    parser.add_argument("--max-pdf-pages", type=int, default=6)
    parser.add_argument("--max-sheet-rows", type=int, default=80)
    args = parser.parse_args()

    started = time.perf_counter()
    downloads_dir = Path(args.downloads_dir)
    onedrive_dir = Path(args.onedrive_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    existing_hashes, existing_titles = load_existing_bundle(Path(args.existing_bundle))
    paths = find_files(downloads_dir, onedrive_dir, datetime.fromisoformat(args.downloads_since))
    candidates = build_candidates(paths, downloads_dir, onedrive_dir, existing_hashes, existing_titles)

    sources: list[SourcePayload] = []
    records: list[CandidateRecord] = []
    items: list[ItemPayload] = []
    success_count = 0
    failure_count = 0

    for candidate in candidates:
        source, source_records, source_items, success = build_records(candidate, args.max_pdf_pages, args.max_sheet_rows)
        sources.append(source)
        records.extend(source_records)
        items.extend(source_items)
        if success:
            success_count += 1
        else:
            failure_count += 1

    elapsed_ms = int((time.perf_counter() - started) * 1000)
    duplicate_count = sum(1 for candidate in candidates if candidate.duplicate_group)
    load_candidate_records = [
        record for record in records
        if record.extracted_text and not record.duplicate_group and not record.needs_manual_review
    ]
    load_candidate_ids = {record.id for record in load_candidate_records}
    review_candidate_records = [record for record in records if record.extracted_text and record.id not in load_candidate_ids]
    run_record = {
        "source_batch": SOURCE_BATCH,
        "source_count": len(sources),
        "item_count": len(items),
        "success_count": success_count,
        "failure_count": failure_count,
        "elapsed_ms": elapsed_ms,
        "report_path": str(out_dir / "validation-report.json"),
        "status": "completed_with_notice" if failure_count else "completed",
        "details": {
            "downloadsDir": str(downloads_dir),
            "onedriveDir": str(onedrive_dir),
            "existingBundle": args.existing_bundle,
            "downloadsSince": args.downloads_since,
            "duplicateCount": duplicate_count,
            "recordCount": len(records),
            "loadCandidateRecordCount": len(load_candidate_records),
            "manualReviewRecordCount": len(review_candidate_records),
            "noDbMutation": True,
        },
    }
    validation_report = {
        "generated_at": now_iso(),
        "db_mutation_applied": False,
        "scan_roots": {
            "downloads": str(downloads_dir),
            "onedrive_industrial_safety": str(onedrive_dir),
        },
        "summary": run_record,
        "counts": {
            "total_files": len(candidates),
            "parse_success_files": success_count,
            "parse_failure_files": failure_count,
            "duplicate_files": duplicate_count,
            "candidate_records": len(records),
            "upsert_item_candidates": len(items),
            "submission_output_records": sum(1 for record in records if record.recommended_use == "submission_output"),
            "internal_knowledge_records": sum(1 for record in records if record.recommended_use == "internal_knowledge_db"),
            "needs_manual_review_records": sum(1 for record in records if record.needs_manual_review),
        },
        "by_document_kind": count_by(record.document_kind for record in records),
        "by_recommended_use": count_by(record.recommended_use for record in records),
        "inventory": [
            {
                **asdict(candidate),
                "path": str(candidate.path),
            }
            for candidate in candidates
        ],
        "records": [asdict(record) for record in records],
        "duplicates": [
            {
                **asdict(candidate),
                "path": str(candidate.path),
            }
            for candidate in candidates
            if candidate.duplicate_group
        ],
        "failures": [asdict(record) for record in records if record.failure_reason],
        "schema_mapping": {
            "reuse_existing_tables": True,
            "safety_reference_sources": {
                "source_file": "source_path",
                "source_type": "source_type/file_format",
                "duplicate fields": "metadata.duplicateGroup / metadata.duplicateReason",
            },
            "safety_reference_items": {
                "document_kind": "item_type",
                "section_title": "title and payload.section_title",
                "extracted_text": "body and payload.extracted_text",
                "field_labels": "payload.field_labels",
                "recommended_use": "payload.recommended_use",
                "confidence": "payload.confidence",
                "needs_manual_review": "payload.needs_manual_review",
            },
            "safety_reference_ingestion_runs": {
                "dry_run_counts": "source_count/item_count/success_count/failure_count/details",
            },
            "migration_candidate": "Optional safety_reference_form_records table only if first-class form QA columns are required.",
        },
    }
    upsert_payload = {
        "sources": [asdict(source) for source in sources],
        "items": [asdict(item) for item in items],
        "form_records_review_model": [asdict(record) for record in records],
        "ingestion_run": run_record,
        "db_mutation_applied": False,
    }

    write_json(out_dir / "validation-report.json", validation_report)
    write_json(out_dir / "upsert-payload.json", upsert_payload)
    (out_dir / "upsert.sql").write_text(build_upsert_sql(sources, items, run_record), encoding="utf-8")
    (out_dir / "migration-candidate.sql").write_text(build_migration_candidate_sql(), encoding="utf-8")
    report_lines = [
        "# Supabase safety ingestion readiness",
        "",
        f"- generated_at: {validation_report['generated_at']}",
        "- db_mutation_applied: false",
        f"- total_files: {len(candidates)}",
        f"- parse_success_files: {success_count}",
        f"- parse_failure_files: {failure_count}",
        f"- duplicate_files: {duplicate_count}",
        f"- candidate_records: {len(records)}",
        f"- upsert_item_candidates: {len(items)}",
        f"- submission_output_records: {validation_report['counts']['submission_output_records']}",
        f"- internal_knowledge_records: {validation_report['counts']['internal_knowledge_records']}",
        "",
        "## Document Kind Counts",
        "",
        *[f"- {key}: {value}" for key, value in validation_report["by_document_kind"].items()],
        "",
        "## Output Files",
        "",
        "- validation-report.json",
        "- upsert-payload.json",
        "- upsert.sql",
        "- migration-candidate.sql",
        "",
        "## Notes",
        "",
        "- Existing `safety_reference_sources`, `safety_reference_items`, `safety_reference_ingestion_runs` are enough for dry-run upsert.",
        "- `migration-candidate.sql` is optional and was not applied.",
        "- Records marked `needs_manual_review=true` should not be bulk-loaded without OCR/layout review.",
    ]
    (out_dir / "report.md").write_text("\n".join(report_lines), encoding="utf-8")
    print(json.dumps({
        "out_dir": str(out_dir),
        "total_files": len(candidates),
        "parse_success_files": success_count,
        "parse_failure_files": failure_count,
        "duplicate_files": duplicate_count,
        "candidate_records": len(records),
        "upsert_item_candidates": len(items),
        "db_mutation_applied": False,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
