from __future__ import annotations

import importlib.util
import inspect
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from types import ModuleType
from unittest import mock

from openpyxl import Workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[1]
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from parser_safety import ParserBudget, ParserLimits
from pdf_parser_worker import PdfWorkerLimitError
from xlsx_parser_worker import XlsxWorkerLimitError


def load_script(filename: str) -> ModuleType:
    script_path = SCRIPTS_DIR / filename
    module_name = f"{script_path.stem}_operator_budget_test"
    spec = importlib.util.spec_from_file_location(module_name, script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {script_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["hazard", "control"])
    sheet.append(["fall", "guardrail"])
    workbook.save(path)
    workbook.close()


class ParseDownloadSafetyFormsBudgetTest(unittest.TestCase):
    def test_structured_archives_are_preflighted_before_parser_initialization(self) -> None:
        parser = load_script("parse_download_safety_forms.py")

        self.assertNotIn("zipfile.ZipFile(path)", inspect.getsource(parser.read_hwpx))
        self.assertNotIn("zipfile.ZipFile(path)", inspect.getsource(parser.read_zip_listing))
        self.assertIn("open_preflighted_archive(path)", inspect.getsource(parser.read_xlsx))
        self.assertIn("parse_xlsx_bytes_bounded", inspect.getsource(parser.read_xlsx))
        self.assertNotIn("load_workbook", inspect.getsource(parser.read_xlsx))

    def test_xlsx_timeout_fails_closed_without_parser_fallback(self) -> None:
        parser = load_script("parse_download_safety_forms.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "timeout.xlsx"
            write_workbook(path)
            with mock.patch.object(
                parser,
                "parse_xlsx_bytes_bounded",
                side_effect=XlsxWorkerLimitError("timeout", "worker deadline exceeded"),
            ):
                result = parser.parse_file(path, max_pdf_pages=1, max_sheet_rows=10)

        self.assertFalse(result.ok)
        self.assertIn("worker deadline exceeded", result.failure_reason or "")

    def test_pdf_replacement_is_rejected_against_admitted_digest(self) -> None:
        parser = load_script("parse_download_safety_forms.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "replaced.pdf"
            path.write_bytes(b"replacement")
            result = parser.parse_file(
                path,
                max_pdf_pages=1,
                max_sheet_rows=10,
                file_sha256="0" * 64,
            )

        self.assertFalse(result.ok)
        self.assertIn("did not match", result.failure_reason or "")

    def test_pdf_timeout_fails_closed_through_disposable_worker(self) -> None:
        parser = load_script("parse_download_safety_forms.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "timeout.pdf"
            path.write_bytes(b"%PDF-timeout")
            with mock.patch.object(
                parser,
                "parse_pdf_file_bounded",
                side_effect=PdfWorkerLimitError("timeout", "worker deadline exceeded"),
            ):
                result = parser.parse_file(path, max_pdf_pages=1, max_sheet_rows=10)

        self.assertFalse(result.ok)
        self.assertIn("worker deadline exceeded", result.failure_reason or "")

    def test_fails_closed_on_byte_text_cell_and_time_limits(self) -> None:
        parser = load_script("parse_download_safety_forms.py")
        with TemporaryDirectory() as temporary_directory:
            csv_path = Path(temporary_directory) / "safety.csv"
            csv_path.write_text("a,b,c\n1,2,3\n", encoding="utf-8")
            cases = {
                "input bytes exceed limit": ParserLimits(max_input_bytes=1),
                "text chars exceed limit": ParserLimits(max_text_chars=1),
                "total cells exceed limit": ParserLimits(max_total_cells=2),
                "elapsed time exceeds limit": ParserLimits(max_elapsed_seconds=0.0),
            }

            for expected_message, limits in cases.items():
                with self.subTest(expected_message=expected_message):
                    result = parser.parse_file(csv_path, max_pdf_pages=1, max_sheet_rows=10, limits=limits)
                    self.assertFalse(result.ok)
                    self.assertEqual(result.item_count, 0)
                    self.assertIn(expected_message, result.failure_reason or "")

    def test_rejects_oversized_candidates_before_hashing(self) -> None:
        parser = load_script("parse_download_safety_forms.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "oversized.csv"
            path.write_bytes(b"oversized")

            with mock.patch.object(parser, "sha256_file", side_effect=AssertionError("must not hash")):
                hashes, _, _, failures = parser.duplicate_maps([path], ParserLimits(max_input_bytes=1))
                result = parser.parse_file(
                    path,
                    max_pdf_pages=1,
                    max_sheet_rows=10,
                    limits=ParserLimits(max_input_bytes=1),
                    admission_failure=failures[path],
                )

        self.assertEqual(hashes, {})
        self.assertIn("input bytes exceed limit", failures[path])
        self.assertFalse(result.ok)
        self.assertIsNone(parser.build_item(result))


class PrepareSupabaseSafetyIngestionBudgetTest(unittest.TestCase):
    def test_structured_archives_are_preflighted_before_parser_initialization(self) -> None:
        preparer = load_script("prepare_supabase_safety_ingestion.py")

        self.assertNotIn("zipfile.ZipFile(path)", inspect.getsource(preparer.read_hwpx_records))
        self.assertNotIn("zipfile.ZipFile(path)", inspect.getsource(preparer.read_zip_records))
        self.assertNotIn("zipfile.ZipFile(path)", inspect.getsource(preparer.read_pptx_records))
        self.assertIn("open_preflighted_archive(path)", inspect.getsource(preparer.read_xlsx_records))
        self.assertIn("parse_xlsx_bytes_bounded", inspect.getsource(preparer.read_xlsx_records))
        self.assertNotIn("load_workbook", inspect.getsource(preparer.read_xlsx_records))

    def test_xlsx_timeout_returns_no_ingestion_records(self) -> None:
        preparer = load_script("prepare_supabase_safety_ingestion.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "timeout.xlsx"
            write_workbook(path)
            with mock.patch.object(
                preparer,
                "parse_xlsx_bytes_bounded",
                side_effect=XlsxWorkerLimitError("timeout", "worker deadline exceeded"),
            ):
                records, parser_name, failure = preparer.read_xlsx_records(
                    path,
                    10,
                    ParserBudget(),
                )

        self.assertEqual(records, [])
        self.assertEqual(parser_name, "openpyxl-worker")
        self.assertIn("worker deadline exceeded", failure or "")

    def test_pdf_replacement_is_rejected_against_candidate_digest(self) -> None:
        preparer = load_script("prepare_supabase_safety_ingestion.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "replaced.pdf"
            path.write_bytes(b"replacement")
            records, parser_name, failure = preparer.read_pdf_records(
                path,
                1,
                ParserBudget(),
                expected_sha256="0" * 64,
            )

        self.assertEqual(records, [])
        self.assertEqual(parser_name, "pypdf")
        self.assertIn("did not match", failure or "")

    def test_pdf_timeout_returns_no_ingestion_records(self) -> None:
        preparer = load_script("prepare_supabase_safety_ingestion.py")
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "timeout.pdf"
            path.write_bytes(b"%PDF-timeout")
            with mock.patch.object(
                preparer,
                "parse_pdf_file_bounded",
                side_effect=PdfWorkerLimitError("timeout", "worker deadline exceeded"),
            ):
                records, parser_name, failure = preparer.read_pdf_records(
                    path,
                    1,
                    ParserBudget(),
                )

        self.assertEqual(records, [])
        self.assertEqual(parser_name, "pypdf")
        self.assertIn("worker deadline exceeded", failure or "")

    def test_returns_no_records_when_operator_budget_is_exceeded(self) -> None:
        preparer = load_script("prepare_supabase_safety_ingestion.py")
        with TemporaryDirectory() as temporary_directory:
            csv_path = Path(temporary_directory) / "safety.csv"
            csv_path.write_text("header,value\n위험,추락\n", encoding="utf-8")

            for expected_message, limits in (
                ("input bytes exceed limit", ParserLimits(max_input_bytes=1)),
                ("text chars exceed limit", ParserLimits(max_text_chars=1)),
                ("total cells exceed limit", ParserLimits(max_total_cells=2)),
                ("elapsed time exceeds limit", ParserLimits(max_elapsed_seconds=0.0)),
            ):
                with self.subTest(expected_message=expected_message):
                    records, _, failure_reason = preparer.extract_records(csv_path, 1, 10, limits)
                    self.assertEqual(records, [])
                    self.assertIn(expected_message, failure_reason or "")

    def test_rejects_oversized_candidate_before_hashing_and_creates_no_item(self) -> None:
        preparer = load_script("prepare_supabase_safety_ingestion.py")
        with TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            path = root / "oversized.csv"
            path.write_bytes(b"oversized")

            with mock.patch.object(preparer, "sha256_file", side_effect=AssertionError("must not hash")):
                candidates = preparer.build_candidates(
                    [path],
                    root,
                    root,
                    set(),
                    set(),
                    ParserLimits(max_input_bytes=1),
                )
            _, records, items, success = preparer.build_records(
                candidates[0],
                1,
                10,
                ParserLimits(max_input_bytes=1),
            )

        self.assertFalse(success)
        self.assertEqual(items, [])
        self.assertEqual(len(records), 1)
        self.assertIn("input bytes exceed limit", records[0].failure_reason or "")


class IngestSafetyReferenceCatalogBudgetTest(unittest.TestCase):
    def test_pdf_timeout_propagates_without_in_process_parser_fallback(self) -> None:
        ingester = load_script("ingest_safety_reference_catalog.py")
        with mock.patch.object(
            ingester,
            "parse_pdf_bytes_bounded",
            side_effect=PdfWorkerLimitError("timeout", "worker deadline exceeded"),
        ):
            with self.assertRaisesRegex(PdfWorkerLimitError, "worker deadline exceeded"):
                ingester.extract_pdf_text_from_bytes(b"%PDF-timeout", 1)

    def test_rejects_pdf_bytes_before_parser_allocation(self) -> None:
        ingester = load_script("ingest_safety_reference_catalog.py")
        budget = ParserBudget(ParserLimits(max_input_bytes=2))

        with self.assertRaisesRegex(ingester.ParserBudgetError, "input bytes exceed limit"):
            ingester.extract_pdf_text_from_bytes(b"not-a-pdf", 1, budget)

    def test_rejects_csv_text_and_elapsed_budgets(self) -> None:
        ingester = load_script("ingest_safety_reference_catalog.py")
        with TemporaryDirectory() as temporary_directory:
            csv_path = Path(temporary_directory) / "catalog.csv"
            csv_path.write_text("column\nvalue\n", encoding="utf-8")

            text_budget = ParserBudget(ParserLimits(max_text_chars=1))
            with self.assertRaisesRegex(ingester.ParserBudgetError, "text chars exceed limit"):
                ingester.read_csv_dicts(csv_path, text_budget)

            time_budget = ParserBudget(ParserLimits(max_elapsed_seconds=0.0))
            with self.assertRaisesRegex(ingester.ParserBudgetError, "elapsed time exceeds limit"):
                ingester.read_csv_dicts(csv_path, time_budget)


if __name__ == "__main__":
    unittest.main()
