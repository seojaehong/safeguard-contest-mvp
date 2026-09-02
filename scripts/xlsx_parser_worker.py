from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from pdf_parser_worker import (
    PdfWorkerError,
    PdfWorkerLimitError,
    _communicate_bounded,
    _spawn_worker,
)


@dataclass(frozen=True)
class XlsxSheetResult:
    name: str
    rows: tuple[tuple[str, ...], ...]


@dataclass(frozen=True)
class XlsxParseResult:
    sheet_count: int
    sheets: tuple[XlsxSheetResult, ...]
    input_sha256: str


class XlsxWorkerError(RuntimeError):
    pass


class XlsxWorkerLimitError(XlsxWorkerError):
    def __init__(self, code: str, message: str | None = None) -> None:
        self.code = code
        super().__init__(message or code)


def _validate_positive_integer(name: str, value: int) -> None:
    if not isinstance(value, int) or isinstance(value, bool) or value <= 0:
        raise ValueError(f"{name} must be a positive integer")


def _validate_limits(
    *,
    max_rows_per_sheet: int,
    max_sheet_count: int,
    max_cells_per_row: int,
    max_total_rows: int,
    max_total_cells: int,
    max_text_chars: int,
    timeout_seconds: float,
    memory_limit_bytes: int,
    max_output_bytes: int,
) -> None:
    for name, value in {
        "max_rows_per_sheet": max_rows_per_sheet,
        "max_sheet_count": max_sheet_count,
        "max_cells_per_row": max_cells_per_row,
        "max_total_rows": max_total_rows,
        "max_total_cells": max_total_cells,
        "max_text_chars": max_text_chars,
        "memory_limit_bytes": memory_limit_bytes,
        "max_output_bytes": max_output_bytes,
    }.items():
        _validate_positive_integer(name, value)
    if not isinstance(timeout_seconds, (int, float)) or isinstance(timeout_seconds, bool):
        raise ValueError("timeout_seconds must be a positive number")
    if timeout_seconds <= 0:
        raise ValueError("timeout_seconds must be greater than zero")


def _worker_command(
    *,
    max_rows_per_sheet: int,
    max_sheet_count: int,
    max_cells_per_row: int,
    max_total_rows: int,
    max_total_cells: int,
    max_text_chars: int,
    memory_limit_bytes: int,
    max_output_bytes: int,
    expected_sha256: str | None,
    test_mode: str | None,
) -> list[str]:
    command = [
        sys.executable,
        str(Path(__file__).resolve()),
        "--worker",
        "--max-rows-per-sheet",
        str(max_rows_per_sheet),
        "--max-sheet-count",
        str(max_sheet_count),
        "--max-cells-per-row",
        str(max_cells_per_row),
        "--max-total-rows",
        str(max_total_rows),
        "--max-total-cells",
        str(max_total_cells),
        "--max-text-chars",
        str(max_text_chars),
        "--memory-limit-bytes",
        str(memory_limit_bytes),
        "--max-output-bytes",
        str(max_output_bytes),
    ]
    if expected_sha256 is not None:
        command.extend(("--expected-sha256", expected_sha256))
    if test_mode is not None:
        command.extend(("--test-mode", test_mode))
    return command


def _decode_result(raw: bytes) -> XlsxParseResult:
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise XlsxWorkerError(f"malformed XLSX worker output: {exc}") from exc
    if not isinstance(payload, dict):
        raise XlsxWorkerError("malformed XLSX worker payload")
    if payload.get("status") == "error":
        error = payload.get("error")
        if not isinstance(error, dict):
            raise XlsxWorkerError("malformed XLSX worker error")
        code = error.get("code")
        message = error.get("message")
        if not isinstance(code, str) or not isinstance(message, str):
            raise XlsxWorkerError("malformed XLSX worker error fields")
        if error.get("kind") == "limit":
            raise XlsxWorkerLimitError(code, message)
        raise XlsxWorkerError(f"{code}: {message}")
    if payload.get("status") != "ok":
        raise XlsxWorkerError("malformed XLSX worker status")

    raw_sheet_count = payload.get("sheet_count")
    raw_sheets = payload.get("sheets")
    input_sha256 = payload.get("input_sha256")
    if (
        not isinstance(raw_sheet_count, int)
        or isinstance(raw_sheet_count, bool)
        or raw_sheet_count < 0
        or not isinstance(raw_sheets, list)
        or not isinstance(input_sha256, str)
        or len(input_sha256) != 64
    ):
        raise XlsxWorkerError("malformed XLSX worker result fields")
    sheets: list[XlsxSheetResult] = []
    for raw_sheet in raw_sheets:
        if not isinstance(raw_sheet, dict):
            raise XlsxWorkerError("malformed XLSX worker sheet")
        name = raw_sheet.get("name")
        raw_rows = raw_sheet.get("rows")
        if not isinstance(name, str) or not isinstance(raw_rows, list):
            raise XlsxWorkerError("malformed XLSX worker sheet fields")
        rows: list[tuple[str, ...]] = []
        for raw_row in raw_rows:
            if not isinstance(raw_row, list) or not all(isinstance(cell, str) for cell in raw_row):
                raise XlsxWorkerError("malformed XLSX worker row")
            rows.append(tuple(raw_row))
        sheets.append(XlsxSheetResult(name=name, rows=tuple(rows)))
    if raw_sheet_count != len(sheets):
        raise XlsxWorkerError("XLSX worker sheet count mismatch")
    return XlsxParseResult(
        sheet_count=raw_sheet_count,
        sheets=tuple(sheets),
        input_sha256=input_sha256,
    )


def parse_xlsx_bytes_bounded(
    data: bytes,
    *,
    max_rows_per_sheet: int,
    max_sheet_count: int,
    max_cells_per_row: int,
    max_total_rows: int,
    max_total_cells: int,
    max_text_chars: int,
    timeout_seconds: float,
    memory_limit_bytes: int = 512 * 1024 * 1024,
    max_output_bytes: int = 32 * 1024 * 1024,
    expected_sha256: str | None = None,
    _test_mode: str | None = None,
) -> XlsxParseResult:
    if not isinstance(data, bytes):
        raise TypeError("data must be bytes")
    _validate_limits(
        max_rows_per_sheet=max_rows_per_sheet,
        max_sheet_count=max_sheet_count,
        max_cells_per_row=max_cells_per_row,
        max_total_rows=max_total_rows,
        max_total_cells=max_total_cells,
        max_text_chars=max_text_chars,
        timeout_seconds=timeout_seconds,
        memory_limit_bytes=memory_limit_bytes,
        max_output_bytes=max_output_bytes,
    )
    deadline = time.monotonic() + float(timeout_seconds)
    command = _worker_command(
        max_rows_per_sheet=max_rows_per_sheet,
        max_sheet_count=max_sheet_count,
        max_cells_per_row=max_cells_per_row,
        max_total_rows=max_total_rows,
        max_total_cells=max_total_cells,
        max_text_chars=max_text_chars,
        memory_limit_bytes=memory_limit_bytes,
        max_output_bytes=max_output_bytes,
        expected_sha256=expected_sha256,
        test_mode=_test_mode,
    )
    worker_environment = dict(os.environ)
    for variable in (
        "OPENBLAS_NUM_THREADS",
        "OMP_NUM_THREADS",
        "MKL_NUM_THREADS",
        "NUMEXPR_NUM_THREADS",
    ):
        worker_environment[variable] = "1"
    try:
        process, job = _spawn_worker(
            command,
            memory_limit_bytes,
            environment=worker_environment,
        )
        raw, _stderr = _communicate_bounded(
            process,
            job,
            data,
            deadline=deadline,
            timeout_seconds=float(timeout_seconds),
            max_output_bytes=max_output_bytes,
        )
    except PdfWorkerLimitError as exc:
        raise XlsxWorkerLimitError(exc.code, str(exc).replace("PDF parser", "XLSX parser")) from exc
    except PdfWorkerError as exc:
        raise XlsxWorkerError(str(exc).replace("PDF parser", "XLSX parser")) from exc
    if time.monotonic() >= deadline:
        raise XlsxWorkerLimitError(
            "timeout",
            f"XLSX parser worker exceeded {timeout_seconds:.3f} seconds",
        )
    return _decode_result(raw)


def _limit_error(code: str, message: str) -> dict[str, object]:
    return {
        "status": "error",
        "error": {"kind": "limit", "code": code, "message": message},
    }


def _parse_workbook(data: bytes, args: argparse.Namespace) -> dict[str, object]:
    from openpyxl import load_workbook

    input_sha256 = hashlib.sha256(data).hexdigest()
    if args.expected_sha256 is not None and input_sha256 != args.expected_sha256:
        return _limit_error(
            "input_digest_mismatch",
            "XLSX input digest did not match the admitted file",
        )
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) > args.max_sheet_count:
            return _limit_error(
                "sheet_count_limit",
                f"XLSX sheet count exceeds limit: {len(workbook.sheetnames)}/{args.max_sheet_count}",
            )
        sheets: list[dict[str, object]] = []
        total_rows = 0
        total_cells = 0
        total_text_chars = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.max_column > args.max_cells_per_row:
                return _limit_error(
                    "row_cells_limit",
                    f"XLSX row cells exceed limit: {sheet.max_column}/{args.max_cells_per_row}",
                )
            rows: list[list[str]] = []
            for row in sheet.iter_rows(
                max_row=args.max_rows_per_sheet,
                max_col=max(1, sheet.max_column),
                values_only=True,
            ):
                total_rows += 1
                total_cells += len(row)
                if total_rows > args.max_total_rows:
                    return _limit_error("total_rows_limit", "XLSX total rows exceed limit")
                if total_cells > args.max_total_cells:
                    return _limit_error("total_cells_limit", "XLSX total cells exceed limit")
                values = [str(cell) for cell in row if cell is not None]
                total_text_chars += sum(len(value) for value in values)
                if total_text_chars > args.max_text_chars:
                    return _limit_error("text_chars_limit", "XLSX text chars exceed limit")
                if values:
                    rows.append(values)
            sheets.append({"name": sheet_name, "rows": rows})
        return {
            "status": "ok",
            "sheet_count": len(sheets),
            "sheets": sheets,
            "input_sha256": input_sha256,
        }
    finally:
        workbook.close()


def _worker_main(args: argparse.Namespace) -> int:
    if os.name == "posix":
        import resource

        resource.setrlimit(
            resource.RLIMIT_AS,
            (args.memory_limit_bytes, args.memory_limit_bytes),
        )
    data = sys.stdin.buffer.read()
    if args.test_mode == "hang":
        while True:
            time.sleep(1.0)
    try:
        payload = _parse_workbook(data, args)
    except MemoryError:
        payload = _limit_error("memory_limit", "XLSX parser worker exhausted memory")
    except Exception as exc:
        payload = {
            "status": "error",
            "error": {
                "kind": "worker",
                "code": "parse_error",
                "message": f"{type(exc).__name__}: {exc}",
            },
        }
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    if len(encoded) > args.max_output_bytes:
        encoded = json.dumps(
            _limit_error("output_bytes_limit", "XLSX worker output exceeds limit"),
            separators=(",", ":"),
        ).encode("utf-8")
    sys.stdout.buffer.write(encoded)
    sys.stdout.buffer.flush()
    return 0


def _argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Disposable bounded XLSX parser worker")
    parser.add_argument("--worker", action="store_true")
    parser.add_argument("--max-rows-per-sheet", type=int)
    parser.add_argument("--max-sheet-count", type=int)
    parser.add_argument("--max-cells-per-row", type=int)
    parser.add_argument("--max-total-rows", type=int)
    parser.add_argument("--max-total-cells", type=int)
    parser.add_argument("--max-text-chars", type=int)
    parser.add_argument("--memory-limit-bytes", type=int)
    parser.add_argument("--max-output-bytes", type=int)
    parser.add_argument("--expected-sha256")
    parser.add_argument("--test-mode", choices=("hang",), help=argparse.SUPPRESS)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = _argument_parser()
    args = parser.parse_args(argv)
    if not args.worker:
        parser.error("--worker is required")
    required = (
        "max_rows_per_sheet",
        "max_sheet_count",
        "max_cells_per_row",
        "max_total_rows",
        "max_total_cells",
        "max_text_chars",
        "memory_limit_bytes",
        "max_output_bytes",
    )
    missing = [name for name in required if getattr(args, name) is None]
    if missing:
        parser.error(f"missing worker limits: {', '.join(missing)}")
    return _worker_main(args)


if __name__ == "__main__":
    raise SystemExit(main())
