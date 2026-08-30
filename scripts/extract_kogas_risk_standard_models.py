from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import sys
import time
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence

import xlrd
from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from archive_safety import BoundedZipReader
from parser_safety import ParserBudget


ROOT_SOURCE_ID = "kogas-risk-standard-models-20240909"
PUBLISHED_AT = "2024-09-09"
AGENCY = "한국가스공사"
PRIMARY_DOCUMENTS = ["위험성평가표", "TBM 브리핑", "작업계획서"]
RISK_KEYWORDS = [
    "추락",
    "떨어짐",
    "넘어짐",
    "전도",
    "충돌",
    "부딪힘",
    "끼임",
    "깔림",
    "감전",
    "화재",
    "폭발",
    "질식",
    "중독",
    "분진",
    "화학",
    "지게차",
    "차량",
    "비계",
    "굴착",
    "밀폐",
    "절단",
    "베임",
    "소음",
    "진동",
]


@dataclass(frozen=True)
class ArchiveMember:
    member_path: str
    file_format: str
    file_size: int
    status: str
    skipped_reason: str | None
    top_level_scope: str
    sheet_count: int
    sheet_names: list[str]
    main_sheet_name: str | None
    workbook_title: str | None
    process_name: str | None
    candidate_row_count: int
    source_id: str | None
    model_item_id: str | None
    profile: dict[str, object]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def log(message: str) -> None:
    print(message, flush=True)


def compact_text(value: object, limit: int = 1000) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit].rstrip()


def canonical_text(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def dedupe_preserve(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        if not item or item in seen:
            continue
        seen.add(item)
        ordered.append(item)
    return ordered


def json_dump(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def normalize_member_path(member_path: str) -> str:
    normalized = member_path.replace("\\", "/").strip()
    try:
        repaired = normalized.encode("cp437").decode("cp949")
    except UnicodeEncodeError:
        repaired = normalized
    except UnicodeDecodeError:
        repaired = normalized
    if repaired != normalized and repaired.count("가") + repaired.count("위") + repaired.count("작") > normalized.count("가") + normalized.count("위") + normalized.count("작"):
        return repaired
    return normalized


def build_source_id(member_path: str) -> str:
    return f"{ROOT_SOURCE_ID}::{member_path}"


def build_item_id(source_id: str, suffix: str) -> str:
    return f"{source_id}#{suffix}"


def infer_top_level_scope(member_path: str) -> str:
    return member_path.split("/", 1)[0]


def infer_tags(*values: str) -> list[str]:
    haystack = " ".join(values)
    return [keyword for keyword in RISK_KEYWORDS if keyword in haystack]


def normalize_header_value(value: str) -> str:
    text = re.sub(r"[\s:/]+", "", value)
    return text.replace("(", "").replace(")", "").lower()


def open_workbook_rows(
    member_name: str,
    payload: bytes,
    budget: ParserBudget,
) -> list[tuple[str, list[list[str]]]]:
    if member_name.lower().endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheets: list[tuple[str, list[list[str]]]] = []
        for sheet_name in workbook.sheetnames:
            budget.start_sheet()
            worksheet = workbook[sheet_name]
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                budget.consume_row(len(row))
                rows.append([compact_text(cell) for cell in row])
            sheets.append((sheet_name, rows))
        return sheets

    workbook = xlrd.open_workbook(file_contents=payload)
    sheets = []
    for sheet_name in workbook.sheet_names():
        budget.start_sheet()
        worksheet = workbook.sheet_by_name(sheet_name)
        rows = []
        for row_index in range(worksheet.nrows):
            budget.consume_row(worksheet.ncols)
            rows.append([
                compact_text(worksheet.cell_value(row_index, column_index))
                for column_index in range(worksheet.ncols)
            ])
        sheets.append((sheet_name, rows))
    return sheets


def score_sheet(rows: Sequence[Sequence[str]]) -> int:
    preview_rows = [" ".join(cell for cell in row if cell) for row in rows[:14]]
    preview_text = " ".join(preview_rows)
    score = 0
    if "위험성평가" in preview_text or "위 험 성 평 가 표" in preview_text:
        score += 6
    if "세부 작업명" in preview_text or "세부작업명" in preview_text:
        score += 5
    if "유해위험요인" in preview_text:
        score += 4
    if "위험성 감소대책" in preview_text:
        score += 4
    if "현재안전보건조치" in preview_text:
        score += 3
    if "관련근거" in preview_text:
        score += 1
    return score


def detect_main_sheet(sheets: Sequence[tuple[str, list[list[str]]]]) -> tuple[str | None, list[list[str]] | None]:
    best_name: str | None = None
    best_rows: list[list[str]] | None = None
    best_score = -1
    for sheet_name, rows in sheets:
        score = score_sheet(rows)
        if score > best_score:
            best_name = sheet_name
            best_rows = rows
            best_score = score
    if best_score <= 0:
        return None, None
    return best_name, best_rows


def detect_process_name(rows: Sequence[Sequence[str]]) -> str | None:
    for row in rows[:10]:
        for index, cell in enumerate(row):
            if "공정명" not in cell:
                continue
            for candidate in row[index + 1:index + 5]:
                normalized_candidate = canonical_text(candidate)
                if normalized_candidate.startswith("평가일시") or normalized_candidate.startswith("평가자"):
                    break
                if candidate:
                    return compact_text(candidate, 200)
            for candidate in row[index + 1:]:
                normalized_candidate = canonical_text(candidate)
                if normalized_candidate.startswith("평가일시") or normalized_candidate.startswith("평가자"):
                    continue
                if candidate:
                    return compact_text(candidate, 200)
    return None


def detect_header_rows(rows: Sequence[Sequence[str]]) -> tuple[int, int] | None:
    max_index = max(0, len(rows) - 1)
    for index in range(min(12, max_index)):
        current_text = canonical_text(" ".join(rows[index]))
        next_text = canonical_text(" ".join(rows[index + 1])) if index + 1 < len(rows) else ""
        combined = current_text + next_text
        if "세부작업명" in current_text and "유해위험요인" in combined and "위험성감소대책" in combined:
            return index, min(index + 1, len(rows) - 1)
    return None


def combine_headers(first: Sequence[str], second: Sequence[str]) -> list[str]:
    width = max(len(first), len(second))
    combined: list[str] = []
    for index in range(width):
        left = first[index] if index < len(first) else ""
        right = second[index] if index < len(second) else ""
        parts = dedupe_preserve([compact_text(left, 120), compact_text(right, 120)])
        combined.append(" ".join(parts))
    return combined


def find_column(headers: Sequence[str], *patterns: str, exclude: Sequence[str] | None = None) -> int | None:
    exclude = exclude or []
    normalized_headers = [normalize_header_value(header) for header in headers]
    for pattern in patterns:
        normalized_pattern = normalize_header_value(pattern)
        for index, header in enumerate(normalized_headers):
            if normalized_pattern in header and not any(normalize_header_value(item) in header for item in exclude):
                return index
    return None


def find_best_column(
    headers: Sequence[str],
    *,
    exact: Sequence[str] | None = None,
    startswith: Sequence[str] | None = None,
    contains: Sequence[str] | None = None,
) -> int | None:
    normalized_headers = [normalize_header_value(header) for header in headers]
    for candidate in exact or []:
        normalized_candidate = normalize_header_value(candidate)
        for index, header in enumerate(normalized_headers):
            if header == normalized_candidate:
                return index
    for candidate in startswith or []:
        normalized_candidate = normalize_header_value(candidate)
        for index, header in enumerate(normalized_headers):
            if header.startswith(normalized_candidate):
                return index
    for candidate in contains or []:
        normalized_candidate = normalize_header_value(candidate)
        for index, header in enumerate(normalized_headers):
            if normalized_candidate in header:
                return index
    return None


def cell(row: Sequence[str], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return compact_text(row[index], 600)


def carry_forward(previous: str | None, current: str) -> tuple[str, bool]:
    if current:
        return current, False
    if previous:
        return previous, True
    return "", False


def choose_title(member_path: str, process_name: str | None) -> str:
    file_stem = Path(member_path).stem
    cleaned = re.sub(r"_r\d+$", "", file_stem, flags=re.IGNORECASE)
    return process_name or compact_text(cleaned, 180)


def build_body_lines(row_record: dict[str, object]) -> str:
    ordered_keys = [
        "process_name",
        "task",
        "hazard_category",
        "hazard_cause",
        "hazard_detail",
        "legal_basis",
        "current_controls",
        "additional_controls",
        "current_risk_likelihood",
        "current_risk_severity",
        "current_risk_score",
        "post_risk_level",
        "owner",
        "due_date",
        "completion_date",
    ]
    lines: list[str] = []
    for key in ordered_keys:
        value = compact_text(row_record.get(key, ""), 800)
        if value:
            lines.append(f"{key}: {value}")
    return "\n".join(lines)


def build_keywords(*values: str) -> list[str]:
    keywords: list[str] = []
    for value in values:
        cleaned = compact_text(value, 120)
        if cleaned:
            keywords.append(cleaned)
    return dedupe_preserve(keywords)


def extract_rows_from_main_sheet(
    member_path: str,
    sheet_name: str,
    rows: Sequence[Sequence[str]],
) -> tuple[str | None, list[dict[str, object]], dict[str, object]]:
    header_rows = detect_header_rows(rows)
    if header_rows is None:
        return None, [], {"reason": "risk-table-header-not-found"}

    header_top_index, header_bottom_index = header_rows
    headers = combine_headers(rows[header_top_index], rows[header_bottom_index])
    task_index = find_best_column(headers, exact=["세부작업명", "세부 작업명"])
    category_index = find_best_column(headers, exact=["유해위험요인파악분류", "분류"], contains=["분류"])
    cause_index = find_best_column(headers, exact=["원인"], contains=["원인"])
    hazard_index = find_best_column(headers, exact=["유해위험요인"], startswith=["유해위험요인"], contains=["유해위험요인"])
    legal_index = find_best_column(headers, exact=["관련근거법규노출기준등"], contains=["관련근거", "법규노출기준등"])
    current_controls_index = find_best_column(headers, exact=["현재안전보건조치"], startswith=["현재안전보건조치"])
    likelihood_index = find_best_column(headers, exact=["현재위험성가능성빈도", "가능성빈도"], contains=["가능성"])
    severity_index = find_best_column(headers, exact=["현재위험성중대성강도", "중대성강도"], contains=["중대성"])
    current_risk_index = find_best_column(headers, exact=["현재위험성위험성", "위험성"])
    additional_controls_index = find_best_column(headers, exact=["위험성감소대책"], startswith=["위험성감소대책"])
    post_risk_index = find_best_column(headers, exact=["개선후위험성등급", "개선후위험성"], startswith=["개선후위험성"])
    due_index = find_best_column(headers, exact=["개선예정일", "개선예정"], contains=["개선예정일", "개선예정"])
    completion_index = find_best_column(headers, exact=["완료일"], contains=["완료일"])
    owner_index = find_best_column(headers, exact=["담당자"], contains=["담당자"])

    process_name = detect_process_name(rows)
    workbook_title = choose_title(member_path, process_name)
    if not process_name:
        process_name = workbook_title
    extracted_rows: list[dict[str, object]] = []
    previous_task: str | None = None
    previous_category: str | None = None
    previous_cause: str | None = None

    for row_number in range(header_bottom_index + 1, len(rows)):
        row = rows[row_number]
        if not any(compact_text(value, 200) for value in row):
            continue

        raw_task = cell(row, task_index)
        raw_category = cell(row, category_index)
        raw_cause = cell(row, cause_index)
        task, task_filled = carry_forward(previous_task, raw_task)
        category, category_filled = carry_forward(previous_category, raw_category)
        cause, cause_filled = carry_forward(previous_cause, raw_cause)
        hazard_detail = cell(row, hazard_index)
        current_controls = cell(row, current_controls_index)
        additional_controls = cell(row, additional_controls_index)
        legal_basis = cell(row, legal_index)
        current_risk_likelihood = cell(row, likelihood_index)
        current_risk_severity = cell(row, severity_index)
        current_risk_score = cell(row, current_risk_index)
        post_risk_level = cell(row, post_risk_index)
        due_date = cell(row, due_index)
        completion_date = cell(row, completion_index)
        owner = cell(row, owner_index)

        if not any([task, category, cause, hazard_detail, current_controls, additional_controls]):
            continue

        if not any([hazard_detail, current_controls, additional_controls]):
            continue

        previous_task = task or previous_task
        previous_category = category or previous_category
        previous_cause = cause or previous_cause

        source_id = build_source_id(member_path)
        row_id = build_item_id(source_id, f"row-{len(extracted_rows) + 1:04d}")
        risk_tags = dedupe_preserve(infer_tags(workbook_title, process_name or "", task, category, cause, hazard_detail, additional_controls))
        row_record: dict[str, object] = {
            "id": row_id,
            "source_id": source_id,
            "root_source_id": ROOT_SOURCE_ID,
            "item_type": "risk-standard-row",
            "archive_member_path": member_path,
            "sheet_name": sheet_name,
            "workbook_title": workbook_title,
            "process_name": process_name,
            "task": task,
            "task_filled_from_previous": task_filled,
            "hazard_category": category,
            "hazard_category_filled_from_previous": category_filled,
            "hazard_cause": cause,
            "hazard_cause_filled_from_previous": cause_filled,
            "hazard_detail": hazard_detail,
            "legal_basis": legal_basis,
            "current_controls": current_controls,
            "additional_controls": additional_controls,
            "current_risk_likelihood": current_risk_likelihood,
            "current_risk_severity": current_risk_severity,
            "current_risk_score": current_risk_score,
            "post_risk_level": post_risk_level,
            "due_date": due_date,
            "completion_date": completion_date,
            "owner": owner,
            "row_index_1based": row_number + 1,
            "row_title": compact_text(" / ".join(part for part in [process_name or workbook_title, task, hazard_detail] if part), 220),
            "keywords": build_keywords(infer_top_level_scope(member_path), process_name or "", task, category, cause),
            "risk_tags": risk_tags,
            "raw_cells": {
                f"column_{index + 1}": compact_text(value, 600)
                for index, value in enumerate(row)
                if compact_text(value, 600)
            },
        }
        row_record["body"] = build_body_lines(row_record)
        extracted_rows.append(row_record)

    extraction_profile = {
        "header_rows_1based": [header_top_index + 1, header_bottom_index + 1],
        "headers": headers,
        "column_map": {
            "task": task_index,
            "hazard_category": category_index,
            "hazard_cause": cause_index,
            "hazard_detail": hazard_index,
            "legal_basis": legal_index,
            "current_controls": current_controls_index,
            "current_risk_likelihood": likelihood_index,
            "current_risk_severity": severity_index,
            "current_risk_score": current_risk_index,
            "additional_controls": additional_controls_index,
            "post_risk_level": post_risk_index,
            "due_date": due_index,
            "completion_date": completion_index,
            "owner": owner_index,
        },
    }
    return workbook_title, extracted_rows, extraction_profile


def build_source_row(member: ArchiveMember) -> dict[str, object]:
    if member.source_id is None:
        raise RuntimeError(f"source_id is missing for parsed member: {member.member_path}")
    return {
        "id": member.source_id,
        "source_group": "kogas-risk-standard-model",
        "source_type": "workbook",
        "agency": AGENCY,
        "title": member.workbook_title or Path(member.member_path).stem,
        "source_path": f"{member.profile['archive_path']}::{member.member_path}",
        "origin_url": None,
        "file_format": member.file_format,
        "published_at": PUBLISHED_AT,
        "metadata": {
            "rootSourceId": ROOT_SOURCE_ID,
            "archivePath": member.profile["archive_path"],
            "memberPath": member.member_path,
            "fileSize": member.file_size,
            "topLevelScope": member.top_level_scope,
            "sheetCount": member.sheet_count,
            "sheetNames": member.sheet_names,
            "mainSheetName": member.main_sheet_name,
            "processName": member.process_name,
            "candidateRowCount": member.candidate_row_count,
            "parser": "openpyxl/xlrd",
            "offlinePreparedAt": now_iso(),
        },
    }


def build_model_item(member: ArchiveMember, row_items: Sequence[dict[str, object]]) -> dict[str, object]:
    if member.source_id is None or member.model_item_id is None:
        raise RuntimeError(f"missing IDs for parsed member: {member.member_path}")
    preview_tasks = dedupe_preserve([compact_text(item.get("task", ""), 80) for item in row_items])[:6]
    preview_hazards = dedupe_preserve([compact_text(item.get("hazard_detail", ""), 100) for item in row_items])[:6]
    aggregated_tags = dedupe_preserve(
        tag
        for item in row_items
        for tag in item.get("risk_tags", [])
        if isinstance(tag, str)
    )
    return {
        "id": member.model_item_id,
        "source_id": member.source_id,
        "item_type": "risk-standard-model",
        "category": member.top_level_scope,
        "subcategory": member.process_name,
        "title": member.workbook_title or Path(member.member_path).stem,
        "summary": compact_text(
            f"{member.process_name or member.workbook_title} 표준모델. {member.candidate_row_count}개 위험성평가 행, {member.sheet_count}개 시트, 메인 시트 {member.main_sheet_name}.",
            220,
        ),
        "body": "\n".join(
            [
                f"member_path: {member.member_path}",
                f"process_name: {member.process_name or ''}",
                f"sheet_names: {', '.join(member.sheet_names)}",
                f"preview_tasks: {', '.join(preview_tasks)}",
                f"preview_hazards: {', '.join(preview_hazards)}",
            ]
        ),
        "keywords": build_keywords(member.top_level_scope, member.process_name or "", *(preview_tasks[:4])),
        "risk_tags": aggregated_tags[:12],
        "primary_documents": PRIMARY_DOCUMENTS,
        "controls": dedupe_preserve([compact_text(item.get("additional_controls", ""), 120) for item in row_items])[:8],
        "payload": {
            "rootSourceId": ROOT_SOURCE_ID,
            "archiveMemberPath": member.member_path,
            "sheetNames": member.sheet_names,
            "mainSheetName": member.main_sheet_name,
            "candidateRowCount": member.candidate_row_count,
            "sampleRowIds": [item["id"] for item in row_items[:5]],
        },
    }


def build_row_item_payload(row: dict[str, object]) -> dict[str, object]:
    return {
        "id": row["id"],
        "source_id": row["source_id"],
        "item_type": "risk-standard-row",
        "category": row["task"] or row["hazard_category"] or row["process_name"],
        "subcategory": row["hazard_category"] or row["hazard_cause"],
        "title": row["row_title"],
        "summary": compact_text(
            " / ".join(
                part
                for part in [
                    compact_text(row.get("task", ""), 80),
                    compact_text(row.get("hazard_detail", ""), 120),
                    compact_text(row.get("additional_controls", ""), 120),
                ]
                if part
            ),
            260,
        ),
        "body": row["body"],
        "keywords": row["keywords"],
        "risk_tags": row["risk_tags"],
        "primary_documents": PRIMARY_DOCUMENTS,
        "controls": dedupe_preserve(
            [
                compact_text(row.get("current_controls", ""), 120),
                compact_text(row.get("additional_controls", ""), 120),
            ]
        ),
        "payload": {
            "rootSourceId": row["root_source_id"],
            "archiveMemberPath": row["archive_member_path"],
            "sheetName": row["sheet_name"],
            "processName": row["process_name"],
            "task": row["task"],
            "hazardCategory": row["hazard_category"],
            "hazardCause": row["hazard_cause"],
            "hazardDetail": row["hazard_detail"],
            "legalBasis": row["legal_basis"],
            "currentControls": row["current_controls"],
            "additionalControls": row["additional_controls"],
            "currentRisk": {
                "likelihood": row["current_risk_likelihood"],
                "severity": row["current_risk_severity"],
                "score": row["current_risk_score"],
            },
            "postRiskLevel": row["post_risk_level"],
            "owner": row["owner"],
            "dueDate": row["due_date"],
            "completionDate": row["completion_date"],
            "rowIndex1Based": row["row_index_1based"],
            "carryForward": {
                "task": row["task_filled_from_previous"],
                "hazardCategory": row["hazard_category_filled_from_previous"],
                "hazardCause": row["hazard_cause_filled_from_previous"],
            },
            "rawCells": row["raw_cells"],
        },
    }


def select_representative_samples(rows: Sequence[dict[str, object]], limit: int) -> list[dict[str, object]]:
    by_scope: dict[str, dict[str, object]] = {}
    ordered: list[dict[str, object]] = []
    for row in rows:
        member_path = str(row["archive_member_path"])
        scope = infer_top_level_scope(member_path)
        if scope not in by_scope:
            by_scope[scope] = row
            ordered.append(row)
        if len(ordered) >= limit:
            return ordered
    return list(rows[:limit])


def render_qa_markdown(
    archive_path: Path,
    profile: dict[str, object],
    representative_samples: Sequence[dict[str, object]],
    output_dir: Path,
) -> str:
    member_counts = profile["member_counts"]
    skipped_reasons = profile["skipped_reasons"]
    lines = [
        "# KOGAS Risk Standard Model Prep QA",
        "",
        f"- Source archive: `{archive_path}`",
        f"- Generated at: `{profile['generated_at']}`",
        f"- Parsed workbook members: `{profile['workbook_counts']['parsed']}`",
        f"- Skipped members: `{profile['workbook_counts']['skipped']}`",
        f"- Extracted risk-standard rows: `{profile['candidate_row_counts']['total']}`",
        "",
        "## Archive profile",
        "",
        f"- Total file members: `{member_counts['total']}`",
        f"- `.xls`: `{member_counts['xls']}`",
        f"- `.xlsx`: `{member_counts['xlsx']}`",
        f"- `.hwp`: `{member_counts['hwp']}`",
        f"- nested `.zip`: `{member_counts['zip']}`",
        f"- Skipped reasons: `{json.dumps(skipped_reasons, ensure_ascii=False)}`",
        "",
        "## Representative extracted samples",
        "",
    ]

    for index, sample in enumerate(representative_samples, start=1):
        lines.extend(
            [
                f"### Sample {index}",
                "",
                f"- Member: `{sample['archive_member_path']}`",
                f"- Process: `{sample.get('process_name', '')}`",
                f"- Task: `{sample.get('task', '')}`",
                f"- Hazard category/cause: `{sample.get('hazard_category', '')}` / `{sample.get('hazard_cause', '')}`",
                f"- Hazard detail: `{sample.get('hazard_detail', '')}`",
                f"- Controls: `{sample.get('additional_controls', '')}`",
                f"- Legal basis: `{sample.get('legal_basis', '')}`",
                f"- Row id: `{sample['id']}`",
                "",
            ]
        )

    lines.extend(
        [
            "## Migration risks",
            "",
            "1. `.hwp` 1건과 nested `.zip` 1건은 이번 오프라인 준비에서 의도적으로 스킵했다. 별도 파서 검토 없이는 동일 규칙으로 섞어 넣으면 안 된다.",
            "2. 메인 표는 대체로 2단 헤더지만, 시트별로 `관련근거`, `현재안전보건조치`, `완료일` 유무가 달라서 DB 적재 전 null 허용 규칙을 다시 확인해야 한다.",
            "3. 일부 행은 `세부 작업명`, `분류`, `원인`이 빈칸이고 윗행 값을 이어받아야 한다. 이번 산출물은 carry-forward 여부를 payload에 남겼으니, 이 플래그를 버리지 말아야 한다.",
            "4. `공정명`과 파일명은 비슷하지만 완전히 같지 않은 경우가 있다. 검색 title은 workbook title, retrieval payload는 process_name과 member path를 함께 보존하는 편이 안전하다.",
            "5. 동일하거나 유사한 감소대책 문구가 여러 템플릿에 반복된다. dedupe는 row ingest 이후 검색/embedding 단계에서 검토하고, raw text는 지금 그대로 보존해야 한다.",
            "",
            "## Offline package artifacts",
            "",
            f"- JSON rows: `{output_dir / 'normalized-risk-standard-rows.json'}`",
            f"- Source profile: `{output_dir / 'source-profile.json'}`",
            f"- Upsert preview: `{output_dir / 'safety-reference-upsert-preview.json'}`",
            f"- Handoff: `{output_dir / 'db-migration-agent-handoff.md'}`",
        ]
    )
    return "\n".join(lines) + "\n"


def build_mapping_proposal(
    archive_path: Path,
    members: Sequence[ArchiveMember],
    row_items: Sequence[dict[str, object]],
    model_items: Sequence[dict[str, object]],
    source_rows: Sequence[dict[str, object]],
) -> dict[str, object]:
    return {
        "root_source_id": ROOT_SOURCE_ID,
        "archive_path": str(archive_path),
        "published_at": PUBLISHED_AT,
        "field_mapping": {
            "safety_reference_sources": {
                "id": "kogas-risk-standard-models-20240909::<archive-member-path>",
                "source_group": "kogas-risk-standard-model",
                "source_type": "workbook",
                "agency": AGENCY,
                "title": "workbook title or process name",
                "source_path": "<archive-path>::<archive-member-path>",
                "origin_url": None,
                "file_format": "xls | xlsx",
                "published_at": PUBLISHED_AT,
                "metadata": {
                    "rootSourceId": ROOT_SOURCE_ID,
                    "memberPath": "<archive-member-path>",
                    "topLevelScope": "<top folder>",
                    "sheetCount": "<int>",
                    "mainSheetName": "<sheet name>",
                    "candidateRowCount": "<int>",
                    "parser": "openpyxl/xlrd",
                },
            },
            "safety_reference_items": {
                "risk-standard-model": {
                    "id": "source_id#model",
                    "source_id": "child workbook source id",
                    "item_type": "risk-standard-model",
                    "category": "top-level scope",
                    "subcategory": "process_name",
                    "title": "workbook title",
                    "summary": "workbook-level summary",
                    "body": "sheet names + preview tasks/hazards",
                    "keywords": "scope + process + preview tasks",
                    "risk_tags": "aggregated hazard tags",
                    "primary_documents": PRIMARY_DOCUMENTS,
                    "controls": "representative additional controls",
                    "payload": {
                        "archiveMemberPath": "<archive-member-path>",
                        "candidateRowCount": "<int>",
                        "sampleRowIds": ["source_id#row-0001"],
                    },
                },
                "risk-standard-row": {
                    "id": "source_id#row-0001",
                    "source_id": "child workbook source id",
                    "item_type": "risk-standard-row",
                    "category": "task or hazard category",
                    "subcategory": "hazard category or cause",
                    "title": "process / task / hazard",
                    "summary": "task + hazard + additional controls",
                    "body": "normalized row fields",
                    "keywords": "scope + process + task + category + cause",
                    "risk_tags": "keyword tags from row text",
                    "primary_documents": PRIMARY_DOCUMENTS,
                    "controls": "current + additional controls",
                    "payload": {
                        "rowIndex1Based": "<int>",
                        "sheetName": "<sheet name>",
                        "carryForward": {
                            "task": "<bool>",
                            "hazardCategory": "<bool>",
                            "hazardCause": "<bool>",
                        },
                        "rawCells": {"column_1": "<text>"},
                    },
                },
            },
        },
        "counts": {
            "parsed_sources": len(source_rows),
            "risk_standard_models": len(model_items),
            "risk_standard_rows": len(row_items),
            "skipped_members": len([member for member in members if member.status == "skipped"]),
        },
        "preview": {
            "sources": list(source_rows[:3]),
            "risk_standard_models": list(model_items[:3]),
            "risk_standard_rows": list(row_items[:5]),
        },
    }


def build_handoff_markdown(
    archive_path: Path,
    output_dir: Path,
    profile: dict[str, object],
    mapping_path: Path,
    upsert_preview_path: Path,
) -> str:
    lines = [
        "# KOGAS Offline Migration Handoff",
        "",
        "This note is for a future DB migration agent. Do not mutate Supabase or run a DB migration from this package.",
        "",
        "## What this package contains",
        "",
        f"- Source archive inspected offline: `{archive_path}`",
        f"- Root source id: `{ROOT_SOURCE_ID}`",
        f"- Parsed workbook members: `{profile['workbook_counts']['parsed']}`",
        f"- Skipped members: `{profile['workbook_counts']['skipped']}`",
        f"- Extracted row candidates: `{profile['candidate_row_counts']['total']}`",
        "",
        "Artifacts:",
        "",
        f"- `normalized-risk-standard-rows.json`: workbook row normalization output",
        f"- `source-profile.json`: per-member profile and skip reasons",
        f"- `safety-reference-upsert-preview.json`: offline preview of `safety_reference_sources/items` rows",
        f"- `mapping-proposal.json`: field mapping contract for `risk-standard-model` and `risk-standard-row`",
        "",
        "## Proposed migration contract",
        "",
        "1. Create one `safety_reference_sources` row per parsable workbook member, not one giant row for the entire ZIP.",
        "2. Create one `risk-standard-model` item per workbook to preserve file-level search and provenance.",
        "3. Create one `risk-standard-row` item per extracted risk row. Keep raw text and carry-forward flags in `payload`.",
        "4. Keep `.hwp` and nested `.zip` members out of the first DB ingest. Review them in a separate parser lane.",
        "5. Treat this package as review material. Do not upload until normalization rules and null-handling are explicitly approved.",
        "",
        "## Suggested future agent steps",
        "",
        "1. Review representative rows in `qa.md` and confirm the row granularity is acceptable for retrieval.",
        "2. Review `mapping-proposal.json` and `safety-reference-upsert-preview.json` with the owning agent before any SQL or REST upload.",
        "3. If approved, feed only the preview payload into an upload script in a separate DB-enabled workstream.",
        "4. Preserve skipped members and reasons in the migration ticket so the second-wave parser scope stays visible.",
        "",
        "## Explicit non-goals for this package",
        "",
        "- No DB migration",
        "- No Supabase mutation",
        "- No `.env` or secret changes",
        "- No archive upload",
        "",
        f"See `{mapping_path.name}` and `{upsert_preview_path.name}` for the exact field contract.",
    ]
    return "\n".join(lines) + "\n"


def run_self_check(
    archive_path: Path,
    output_dir: Path,
    profile_path: Path,
    rows_path: Path,
    mapping_path: Path,
    upsert_preview_path: Path,
) -> dict[str, object]:
    issues: list[str] = []
    profile = json.loads(profile_path.read_text(encoding="utf-8"))
    rows_payload = json.loads(rows_path.read_text(encoding="utf-8"))
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    upsert_preview = json.loads(upsert_preview_path.read_text(encoding="utf-8"))

    expected_artifacts = [
        output_dir / "qa.md",
        output_dir / "report.json",
        output_dir / "db-migration-agent-handoff.md",
        output_dir / "mapping-proposal.json",
        output_dir / "normalized-risk-standard-rows.json",
        output_dir / "safety-reference-upsert-preview.json",
        output_dir / "source-profile.json",
    ]
    for artifact in expected_artifacts:
        if not artifact.exists():
            issues.append(f"missing artifact: {artifact.name}")

    with zipfile.ZipFile(archive_path) as archive:
        file_members = [normalize_member_path(member.filename) for member in archive.infolist() if not member.is_dir()]

    if profile["member_counts"]["total"] != len(file_members):
        issues.append("member count mismatch against archive")

    skipped_reasons = profile.get("skipped_reasons", {})
    if skipped_reasons.get("unsupported-hwp") != 1:
        issues.append("expected exactly one unsupported-hwp skip")
    if skipped_reasons.get("nested-zip-offline-review-required") != 1:
        issues.append("expected exactly one nested-zip skip")

    if rows_payload.get("row_count", 0) <= 0:
        issues.append("normalized row output is empty")

    if mapping.get("counts", {}).get("risk_standard_rows") != rows_payload.get("row_count"):
        issues.append("mapping row count does not match normalized rows")

    source_count = len(upsert_preview.get("sources", []))
    model_count = len([item for item in upsert_preview.get("items", []) if item.get("item_type") == "risk-standard-model"])
    row_count = len([item for item in upsert_preview.get("items", []) if item.get("item_type") == "risk-standard-row"])
    if source_count != model_count:
        issues.append("source count and model item count should match")
    if row_count != rows_payload.get("row_count"):
        issues.append("upsert preview row count does not match normalized rows")

    verdict = "pass" if not issues else "fail"
    result = {
        "generated_at": now_iso(),
        "verdict": verdict,
        "issue_count": len(issues),
        "issues": issues,
        "summary": {
            "archive_members": len(file_members),
            "parsed_sources": source_count,
            "risk_standard_models": model_count,
            "risk_standard_rows": row_count,
        },
    }
    json_dump(output_dir / "self-check.json", result)
    return result


def generate_package(archive_path: Path, output_dir: Path, sample_limit: int) -> dict[str, object]:
    started_at = time.perf_counter()
    generated_at = now_iso()
    output_dir.mkdir(parents=True, exist_ok=True)

    members: list[ArchiveMember] = []
    normalized_rows: list[dict[str, object]] = []
    source_rows: list[dict[str, object]] = []
    model_items: list[dict[str, object]] = []
    row_items: list[dict[str, object]] = []
    skipped_reasons: dict[str, int] = {}
    file_format_counts = {"xls": 0, "xlsx": 0, "hwp": 0, "zip": 0}
    total_sheet_count = 0

    parser_budget = ParserBudget()
    parser_budget.assert_input_file(archive_path)
    with zipfile.ZipFile(archive_path) as archive:
        bounded_archive = BoundedZipReader(archive)
        file_infos = [info for info in bounded_archive.infos if not info.is_dir()]
        for info in file_infos:
            member_path = normalize_member_path(info.filename)
            suffix = Path(member_path).suffix.lower()
            top_level_scope = infer_top_level_scope(member_path)

            if suffix == ".hwp":
                skipped_reasons["unsupported-hwp"] = skipped_reasons.get("unsupported-hwp", 0) + 1
                file_format_counts["hwp"] += 1
                members.append(
                    ArchiveMember(
                        member_path=member_path,
                        file_format="hwp",
                        file_size=info.file_size,
                        status="skipped",
                        skipped_reason="unsupported-hwp",
                        top_level_scope=top_level_scope,
                        sheet_count=0,
                        sheet_names=[],
                        main_sheet_name=None,
                        workbook_title=None,
                        process_name=None,
                        candidate_row_count=0,
                        source_id=None,
                        model_item_id=None,
                        profile={"archive_path": str(archive_path)},
                    )
                )
                continue

            if suffix == ".zip":
                skipped_reasons["nested-zip-offline-review-required"] = skipped_reasons.get("nested-zip-offline-review-required", 0) + 1
                file_format_counts["zip"] += 1
                members.append(
                    ArchiveMember(
                        member_path=member_path,
                        file_format="zip",
                        file_size=info.file_size,
                        status="skipped",
                        skipped_reason="nested-zip-offline-review-required",
                        top_level_scope=top_level_scope,
                        sheet_count=0,
                        sheet_names=[],
                        main_sheet_name=None,
                        workbook_title=None,
                        process_name=None,
                        candidate_row_count=0,
                        source_id=None,
                        model_item_id=None,
                        profile={"archive_path": str(archive_path)},
                    )
                )
                continue

            if suffix not in {".xls", ".xlsx"}:
                skipped_reasons[f"unsupported-{suffix or 'no-extension'}"] = skipped_reasons.get(f"unsupported-{suffix or 'no-extension'}", 0) + 1
                members.append(
                    ArchiveMember(
                        member_path=member_path,
                        file_format=suffix.lstrip("."),
                        file_size=info.file_size,
                        status="skipped",
                        skipped_reason=f"unsupported-{suffix or 'no-extension'}",
                        top_level_scope=top_level_scope,
                        sheet_count=0,
                        sheet_names=[],
                        main_sheet_name=None,
                        workbook_title=None,
                        process_name=None,
                        candidate_row_count=0,
                        source_id=None,
                        model_item_id=None,
                        profile={"archive_path": str(archive_path)},
                    )
                )
                continue

            file_format_counts[suffix.lstrip(".")] += 1
            payload = bounded_archive.read(info)
            sheets = open_workbook_rows(member_path, payload, parser_budget)
            total_sheet_count += len(sheets)
            main_sheet_name, main_sheet_rows = detect_main_sheet(sheets)
            process_name = detect_process_name(main_sheet_rows or [])
            workbook_title: str | None = None
            extraction_profile: dict[str, object]
            extracted_rows: list[dict[str, object]]

            if main_sheet_name is None or main_sheet_rows is None:
                skipped_reasons["no-risk-table-detected"] = skipped_reasons.get("no-risk-table-detected", 0) + 1
                members.append(
                    ArchiveMember(
                        member_path=member_path,
                        file_format=suffix.lstrip("."),
                        file_size=info.file_size,
                        status="skipped",
                        skipped_reason="no-risk-table-detected",
                        top_level_scope=top_level_scope,
                        sheet_count=len(sheets),
                        sheet_names=[sheet_name for sheet_name, _ in sheets],
                        main_sheet_name=None,
                        workbook_title=None,
                        process_name=process_name,
                        candidate_row_count=0,
                        source_id=None,
                        model_item_id=None,
                        profile={"archive_path": str(archive_path)},
                    )
                )
                continue

            workbook_title, extracted_rows, extraction_profile = extract_rows_from_main_sheet(member_path, main_sheet_name, main_sheet_rows)
            source_id = build_source_id(member_path)
            model_item_id = build_item_id(source_id, "model")
            member = ArchiveMember(
                member_path=member_path,
                file_format=suffix.lstrip("."),
                file_size=info.file_size,
                status="parsed",
                skipped_reason=None,
                top_level_scope=top_level_scope,
                sheet_count=len(sheets),
                sheet_names=[sheet_name for sheet_name, _ in sheets],
                main_sheet_name=main_sheet_name,
                workbook_title=workbook_title,
                process_name=process_name,
                candidate_row_count=len(extracted_rows),
                source_id=source_id,
                model_item_id=model_item_id,
                profile={
                    "archive_path": str(archive_path),
                    "extraction_profile": extraction_profile,
                },
            )
            members.append(member)
            normalized_rows.extend(extracted_rows)
            source_rows.append(build_source_row(member))
            model_items.append(build_model_item(member, extracted_rows))
            row_items.extend(build_row_item_payload(row) for row in extracted_rows)

    member_dicts: list[dict[str, object]] = []
    for member in members:
        data = asdict(member)
        if member.status == "parsed":
            extraction_profile = member.profile.get("extraction_profile", {})
            if isinstance(extraction_profile, dict):
                data["extraction_profile"] = extraction_profile
        member_dicts.append(data)

    profile = {
        "generated_at": generated_at,
        "archive_path": str(archive_path),
        "root_source_id": ROOT_SOURCE_ID,
        "member_counts": {
            "total": len(members),
            "xls": file_format_counts["xls"],
            "xlsx": file_format_counts["xlsx"],
            "hwp": file_format_counts["hwp"],
            "zip": file_format_counts["zip"],
        },
        "workbook_counts": {
            "parsed": len([member for member in members if member.status == "parsed"]),
            "skipped": len([member for member in members if member.status == "skipped"]),
        },
        "sheet_counts": {
            "total": total_sheet_count,
        },
        "candidate_row_counts": {
            "total": len(normalized_rows),
            "max_per_member": max((member.candidate_row_count for member in members), default=0),
        },
        "skipped_reasons": skipped_reasons,
        "members": member_dicts,
    }

    mapping_proposal = build_mapping_proposal(archive_path, members, row_items, model_items, source_rows)
    representative_samples = select_representative_samples(normalized_rows, sample_limit)

    rows_payload = {
        "generated_at": generated_at,
        "archive_path": str(archive_path),
        "root_source_id": ROOT_SOURCE_ID,
        "row_count": len(normalized_rows),
        "rows": normalized_rows,
    }
    upsert_preview = {
        "generated_at": generated_at,
        "root_source_id": ROOT_SOURCE_ID,
        "sources": source_rows,
        "items": [*model_items, *row_items],
    }

    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
    report = {
        "generated_at": generated_at,
        "archive_path": str(archive_path),
        "root_source_id": ROOT_SOURCE_ID,
        "member_count": len(members),
        "success_count": len([member for member in members if member.status == "parsed"]),
        "failure_count": len([member for member in members if member.status == "skipped"]),
        "sheet_count": total_sheet_count,
        "row_count": len(normalized_rows),
        "elapsed_ms": elapsed_ms,
        "artifacts": [
            "db-migration-agent-handoff.md",
            "mapping-proposal.json",
            "normalized-risk-standard-rows.json",
            "qa.md",
            "report.json",
            "safety-reference-upsert-preview.json",
            "source-profile.json",
        ],
        "skipped_reasons": skipped_reasons,
    }

    profile_path = output_dir / "source-profile.json"
    rows_path = output_dir / "normalized-risk-standard-rows.json"
    mapping_path = output_dir / "mapping-proposal.json"
    upsert_preview_path = output_dir / "safety-reference-upsert-preview.json"
    qa_path = output_dir / "qa.md"
    handoff_path = output_dir / "db-migration-agent-handoff.md"
    report_path = output_dir / "report.json"

    json_dump(profile_path, profile)
    json_dump(rows_path, rows_payload)
    json_dump(mapping_path, mapping_proposal)
    json_dump(upsert_preview_path, upsert_preview)
    qa_path.write_text(render_qa_markdown(archive_path, profile, representative_samples, output_dir), encoding="utf-8")
    handoff_path.write_text(
        build_handoff_markdown(archive_path, output_dir, profile, mapping_path, upsert_preview_path),
        encoding="utf-8",
    )
    json_dump(report_path, report)

    self_check = run_self_check(
        archive_path=archive_path,
        output_dir=output_dir,
        profile_path=profile_path,
        rows_path=rows_path,
        mapping_path=mapping_path,
        upsert_preview_path=upsert_preview_path,
    )
    report["self_check"] = self_check
    json_dump(report_path, report)

    return {
        "report": report,
        "self_check": self_check,
        "output_dir": str(output_dir),
    }


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare an offline KOGAS risk standard model migration package.")
    parser.add_argument(
        "--archive",
        default=r"C:\Users\iceam\Downloads\한국가스공사_KOGAS 위험성평가 표준모델_20240909.zip",
        help="Path to the source KOGAS ZIP archive.",
    )
    parser.add_argument(
        "--output-dir",
        default=r"C:\Users\iceam\dev\safeguard-contest-mvp\.worktrees\backend-harness-gate\evaluation\northstar-72h-2026-07-10\kogas-risk-standard-model-prep",
        help="Directory where offline artifacts should be written.",
    )
    parser.add_argument("--sample-limit", type=int, default=5, help="Representative sample count for QA markdown.")
    parser.add_argument(
        "--self-check-only",
        action="store_true",
        help="Validate existing artifacts in the output directory without regenerating them.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    archive_path = Path(args.archive)
    output_dir = Path(args.output_dir)

    if not archive_path.exists():
        log(f"[error] archive not found: {archive_path}")
        return 1

    profile_path = output_dir / "source-profile.json"
    rows_path = output_dir / "normalized-risk-standard-rows.json"
    mapping_path = output_dir / "mapping-proposal.json"
    upsert_preview_path = output_dir / "safety-reference-upsert-preview.json"

    try:
        if args.self_check_only:
            for required in [profile_path, rows_path, mapping_path, upsert_preview_path]:
                if not required.exists():
                    raise FileNotFoundError(f"self-check requires existing artifact: {required}")
            result = run_self_check(
                archive_path=archive_path,
                output_dir=output_dir,
                profile_path=profile_path,
                rows_path=rows_path,
                mapping_path=mapping_path,
                upsert_preview_path=upsert_preview_path,
            )
            log(json.dumps(result, ensure_ascii=False, indent=2))
            return 0 if result["verdict"] == "pass" else 1

        result = generate_package(
            archive_path=archive_path,
            output_dir=output_dir,
            sample_limit=args.sample_limit,
        )
        log(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if result["self_check"]["verdict"] == "pass" else 1
    except Exception as exc:  # noqa: BLE001
        log(f"[error] {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
