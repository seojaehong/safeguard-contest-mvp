from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import time
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {
    ".csv",
    ".hwp",
    ".hwpx",
    ".pdf",
    ".xls",
    ".xlsx",
    ".zip",
    ".jpg",
    ".jpeg",
    ".png",
}

RISK_KEYWORDS = [
    "TBM",
    "위험성평가",
    "작업계획",
    "안전교육",
    "추락",
    "끼임",
    "충돌",
    "전도",
    "화재",
    "폭발",
    "질식",
    "중독",
    "감전",
    "붕괴",
    "한랭",
    "폭염",
    "장마",
    "해빙기",
    "기계",
    "장비",
    "비계",
    "데크플레이트",
    "산업안전보건관리비",
]

FORM_HINTS = {
    "tbm": ["TBM", "Tool Box", "작업 전 안전점검회의", "회의록", "일지"],
    "risk_assessment": ["위험성평가", "위험요인", "자율점검표"],
    "work_plan": ["작업계획서", "표준 작업계획서", "기계장비"],
    "permit_or_checklist": ["점검표", "자체점검", "안전수칙", "이행각서", "지정서", "문진표"],
    "education": ["교육", "가이드", "길잡이", "교안"],
    "emergency": ["비상상황", "대비", "응급"],
}


@dataclass(frozen=True)
class InventoryEntry:
    id: str
    path: str
    name: str
    extension: str
    size_bytes: int
    modified_at: str
    sha256: str
    duplicate_group: str | None
    duplicate_reason: str | None


@dataclass(frozen=True)
class ParseResult:
    source_id: str
    path: str
    name: str
    extension: str
    parser: str
    ok: bool
    failure_reason: str | None
    title: str
    text_length: int
    page_count: int | None
    sheet_count: int | None
    image_count: int | None
    zip_entry_count: int | None
    item_count: int
    keywords: list[str]
    form_type: str
    sample_text: str


@dataclass(frozen=True)
class SourcePayload:
    id: str
    source_group: str
    source_type: str
    agency: str
    title: str
    source_path: str
    origin_url: str | None
    file_format: str
    published_at: str | None
    metadata: dict[str, Any]


@dataclass(frozen=True)
class ItemPayload:
    id: str
    source_id: str
    item_type: str
    category: str | None
    subcategory: str | None
    title: str
    summary: str
    body: str
    keywords: list[str]
    risk_tags: list[str]
    primary_documents: list[str]
    controls: list[str]
    payload: dict[str, Any]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def slugify(value: str, max_length: int = 100) -> str:
    normalized = re.sub(r"[^0-9A-Za-z가-힣]+", "-", value.strip()).strip("-").lower()
    if not normalized:
        return hashlib.sha1(value.encode("utf-8", errors="ignore")).hexdigest()[:12]
    if len(normalized) <= max_length:
        return normalized
    digest = hashlib.sha1(value.encode("utf-8", errors="ignore")).hexdigest()[:10]
    return f"{normalized[:max_length - 11].rstrip('-')}-{digest}"


def compact_text(value: str, limit: int = 1200) -> str:
    return re.sub(r"\s+", " ", value or "").strip()[:limit].rstrip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_pdf(path: Path, max_pages: int) -> tuple[str, int]:
    reader = PdfReader(str(path))
    texts: list[str] = []
    for page in reader.pages[:max_pages]:
        texts.append(page.extract_text() or "")
    return "\n".join(texts), len(reader.pages)


def read_hwpx(path: Path) -> tuple[str, int, int]:
    texts: list[str] = []
    image_count = 0
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        for name in names:
            lower_name = name.lower()
            if lower_name.endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif")):
                image_count += 1
            if not lower_name.endswith(".xml"):
                continue
            if not ("/section" in lower_name or lower_name.endswith("content.hpf") or "header" in lower_name):
                continue
            try:
                root = ElementTree.fromstring(archive.read(name))
            except ElementTree.ParseError:
                continue
            for element in root.iter():
                if element.text and element.text.strip():
                    texts.append(element.text.strip())
    return "\n".join(texts), len(names), image_count


def read_xlsx(path: Path, max_rows: int) -> tuple[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines.append(f"[sheet] {sheet_name}")
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index > max_rows:
                break
            values = [compact_text(str(cell), 160) for cell in row if cell is not None and compact_text(str(cell), 160)]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines), len(workbook.sheetnames)


def read_csv_text(path: Path, max_rows: int) -> tuple[str, str]:
    encodings = ["utf-8-sig", "cp949", "euc-kr"]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                reader = csv.reader(file)
                rows = [" | ".join(row) for _, row in zip(range(max_rows), reader)]
            return "\n".join(rows), encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"CSV encoding detection failed: {last_error}")


def read_zip_listing(path: Path) -> tuple[str, int]:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
    return "\n".join(names), len(names)


def infer_keywords(name: str, text: str) -> list[str]:
    combined = f"{name}\n{text}"
    found = [keyword for keyword in RISK_KEYWORDS if keyword.lower() in combined.lower()]
    if "Tool Box" in combined and "TBM" not in found:
        found.append("TBM")
    return sorted(set(found), key=found.index)


def infer_form_type(name: str, text: str) -> str:
    combined = f"{name}\n{text}"
    for form_type, hints in FORM_HINTS.items():
        if any(hint.lower() in combined.lower() for hint in hints):
            return form_type
    return "reference_material"


def infer_primary_documents(form_type: str, keywords: list[str]) -> list[str]:
    documents: list[str] = []
    if form_type in {"risk_assessment", "work_plan", "permit_or_checklist"}:
        documents.extend(["위험성평가표", "작업계획서", "허가서/첨부"])
    if form_type in {"tbm", "education"} or "TBM" in keywords:
        documents.extend(["TBM 브리핑", "TBM 기록", "안전보건교육 기록"])
    if form_type == "emergency":
        documents.extend(["비상대응 절차", "사진/증빙"])
    if not documents:
        documents.extend(["위험성평가표", "TBM 브리핑", "안전보건교육 기록"])
    return sorted(set(documents), key=documents.index)


def infer_controls(keywords: list[str], form_type: str) -> list[str]:
    controls: list[str] = []
    if any(keyword in keywords for keyword in ["추락", "비계"]):
        controls.extend(["작업발판·난간·개구부 확인", "안전대 체결 및 출입통제"])
    if any(keyword in keywords for keyword in ["기계", "장비", "충돌", "끼임"]):
        controls.extend(["장비 동선과 보행자 동선 분리", "신호수 지정 및 작업반경 통제"])
    if any(keyword in keywords for keyword in ["화재", "폭발"]):
        controls.extend(["화기작업 허가 및 소화설비 확인", "인화성 물질 격리"])
    if any(keyword in keywords for keyword in ["질식", "중독"]):
        controls.extend(["작업 전 산소·유해가스 측정", "감시인 배치 및 구조장비 확보"])
    if any(keyword in keywords for keyword in ["한랭", "폭염", "장마", "해빙기"]):
        controls.extend(["기상특보 확인", "작업중지 기준과 휴식계획 공유"])
    if form_type == "tbm":
        controls.append("작업 전 TBM 참석자 서명과 위험요인 복창 확인")
    if not controls:
        controls.append("작업 전 유해·위험요인 확인 및 관리감독자 승인")
    return sorted(set(controls), key=controls.index)


def parse_file(path: Path, max_pdf_pages: int, max_sheet_rows: int) -> ParseResult:
    extension = path.suffix.lower()
    parser = "metadata-only"
    text = ""
    page_count: int | None = None
    sheet_count: int | None = None
    image_count: int | None = None
    zip_entry_count: int | None = None
    failure_reason: str | None = None

    try:
        if extension == ".pdf":
            parser = "pypdf"
            text, page_count = read_pdf(path, max_pdf_pages)
        elif extension == ".hwpx":
            parser = "zip+xml"
            text, zip_entry_count, image_count = read_hwpx(path)
        elif extension == ".xlsx":
            parser = "openpyxl"
            text, sheet_count = read_xlsx(path, max_sheet_rows)
        elif extension == ".csv":
            parser = "csv"
            text, encoding = read_csv_text(path, max_sheet_rows)
            parser = f"csv:{encoding}"
        elif extension == ".zip":
            parser = "zip-listing"
            text, zip_entry_count = read_zip_listing(path)
        elif extension in {".jpg", ".jpeg", ".png"}:
            parser = "image-metadata"
            text = f"Image evidence file: {path.name}"
            image_count = 1
        elif extension in {".hwp", ".xls"}:
            failure_reason = f"{extension} binary parser not enabled in this dry-run bundle"
        else:
            failure_reason = f"unsupported extension: {extension}"
    except Exception as exc:
        failure_reason = str(exc)

    ok = bool(text.strip()) and failure_reason is None
    if not ok and failure_reason is None:
        failure_reason = "no extractable text"
    keywords = infer_keywords(path.name, text)
    form_type = infer_form_type(path.name, text)
    source_id = f"downloads-{slugify(path.stem)}-{sha256_file(path)[:10]}"
    return ParseResult(
        source_id=source_id,
        path=str(path),
        name=path.name,
        extension=extension,
        parser=parser,
        ok=ok,
        failure_reason=failure_reason,
        title=path.stem,
        text_length=len(text),
        page_count=page_count,
        sheet_count=sheet_count,
        image_count=image_count,
        zip_entry_count=zip_entry_count,
        item_count=1 if ok else 0,
        keywords=keywords,
        form_type=form_type,
        sample_text=compact_text(text, 1200),
    )


def build_source(path: Path, entry: InventoryEntry, result: ParseResult) -> SourcePayload:
    return SourcePayload(
        id=result.source_id,
        source_group="downloads-safety-forms-2026-05-03",
        source_type=result.form_type,
        agency="KOSHA/MOEL or uploaded form",
        title=result.title,
        source_path=str(path),
        origin_url=None,
        file_format=entry.extension.lstrip("."),
        published_at=None,
        metadata={
            "fileName": entry.name,
            "sizeBytes": entry.size_bytes,
            "sha256": entry.sha256,
            "modifiedAt": entry.modified_at,
            "parser": result.parser,
            "duplicateGroup": entry.duplicate_group,
            "duplicateReason": entry.duplicate_reason,
            "pageCount": result.page_count,
            "sheetCount": result.sheet_count,
            "imageCount": result.image_count,
            "zipEntryCount": result.zip_entry_count,
        },
    )


def build_item(result: ParseResult) -> ItemPayload | None:
    if not result.ok:
        return None
    keywords = result.keywords or [result.form_type]
    documents = infer_primary_documents(result.form_type, keywords)
    return ItemPayload(
        id=f"{result.source_id}-item-0001",
        source_id=result.source_id,
        item_type=result.form_type,
        category=result.form_type,
        subcategory=result.extension.lstrip("."),
        title=result.title,
        summary=compact_text(result.sample_text, 280),
        body=result.sample_text,
        keywords=keywords,
        risk_tags=[keyword for keyword in keywords if keyword not in {"TBM", "위험성평가", "작업계획", "안전교육"}],
        primary_documents=documents,
        controls=infer_controls(keywords, result.form_type),
        payload={
            "sourcePath": result.path,
            "parser": result.parser,
            "textLength": result.text_length,
            "sampleOnly": True,
            "fullChunkingRecommended": result.extension == ".pdf" and (result.page_count or 0) > 20,
        },
    )


def sql_literal(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, list):
        return "array[" + ",".join(sql_literal(item) for item in value) + "]::text[]"
    if isinstance(value, dict):
        return "'" + json.dumps(value, ensure_ascii=False).replace("'", "''") + "'::jsonb"
    return "'" + str(value).replace("'", "''") + "'"


def build_upsert_sql(sources: list[SourcePayload], items: list[ItemPayload], run_record: dict[str, Any]) -> str:
    source_values = ",\n".join(
        "(" + ",".join([
            sql_literal(source.id),
            sql_literal(source.source_group),
            sql_literal(source.source_type),
            sql_literal(source.agency),
            sql_literal(source.title),
            sql_literal(source.source_path),
            sql_literal(source.origin_url),
            sql_literal(source.file_format),
            sql_literal(source.published_at),
            sql_literal(source.metadata),
        ]) + ")"
        for source in sources
    )
    item_values = ",\n".join(
        "(" + ",".join([
            sql_literal(item.id),
            sql_literal(item.source_id),
            sql_literal(item.item_type),
            sql_literal(item.category),
            sql_literal(item.subcategory),
            sql_literal(item.title),
            sql_literal(item.summary),
            sql_literal(item.body),
            sql_literal(item.keywords),
            sql_literal(item.risk_tags),
            sql_literal(item.primary_documents),
            sql_literal(item.controls),
            sql_literal(item.payload),
        ]) + ")"
        for item in items
    )
    parts = [
        "-- Dry-run upsert SQL generated for review only. Do not execute before approval.",
        "insert into safety_reference_sources (id, source_group, source_type, agency, title, source_path, origin_url, file_format, published_at, metadata)",
        f"values\n{source_values}",
        "on conflict (id) do update set",
        "  source_group = excluded.source_group,",
        "  source_type = excluded.source_type,",
        "  agency = excluded.agency,",
        "  title = excluded.title,",
        "  source_path = excluded.source_path,",
        "  origin_url = excluded.origin_url,",
        "  file_format = excluded.file_format,",
        "  published_at = excluded.published_at,",
        "  metadata = excluded.metadata,",
        "  updated_at = now();",
    ]
    if item_values:
        parts.extend([
            "",
            "insert into safety_reference_items (id, source_id, item_type, category, subcategory, title, summary, body, keywords, risk_tags, primary_documents, controls, payload)",
            f"values\n{item_values}",
            "on conflict (id) do update set",
            "  source_id = excluded.source_id,",
            "  item_type = excluded.item_type,",
            "  category = excluded.category,",
            "  subcategory = excluded.subcategory,",
            "  title = excluded.title,",
            "  summary = excluded.summary,",
            "  body = excluded.body,",
            "  keywords = excluded.keywords,",
            "  risk_tags = excluded.risk_tags,",
            "  primary_documents = excluded.primary_documents,",
            "  controls = excluded.controls,",
            "  payload = excluded.payload,",
            "  updated_at = now();",
        ])
    parts.extend([
        "",
        "insert into safety_reference_ingestion_runs (source_batch, source_count, item_count, success_count, failure_count, elapsed_ms, report_path, status, details)",
        "values (" + ",".join([
            sql_literal(run_record["source_batch"]),
            str(run_record["source_count"]),
            str(run_record["item_count"]),
            str(run_record["success_count"]),
            str(run_record["failure_count"]),
            str(run_record["elapsed_ms"]),
            sql_literal(run_record["report_path"]),
            sql_literal(run_record["status"]),
            sql_literal(run_record["details"]),
        ]) + ");",
    ])
    return "\n".join(parts) + "\n"


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def find_target_files(downloads_dir: Path, since: datetime | None) -> list[Path]:
    files: list[Path] = []
    for path in downloads_dir.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if path.name.lower().startswith("client_secret"):
            continue
        modified = datetime.fromtimestamp(path.stat().st_mtime)
        if since and modified < since:
            continue
        files.append(path)
    return sorted(files, key=lambda item: item.stat().st_mtime)


def duplicate_maps(paths: Iterable[Path]) -> tuple[dict[Path, str], dict[Path, str | None], dict[Path, str | None]]:
    hashes = {path: sha256_file(path) for path in paths}
    hash_groups: dict[str, list[Path]] = {}
    stem_groups: dict[str, list[Path]] = {}
    for path, digest in hashes.items():
        hash_groups.setdefault(digest, []).append(path)
        normalized_stem = re.sub(r"\s+", " ", path.stem.replace("_", " ")).strip().lower()
        stem_groups.setdefault(normalized_stem, []).append(path)
    duplicate_group: dict[Path, str | None] = {path: None for path in hashes}
    duplicate_reason: dict[Path, str | None] = {path: None for path in hashes}
    for digest, group in hash_groups.items():
        if len(group) > 1:
            for path in group:
                duplicate_group[path] = f"sha256:{digest[:12]}"
                duplicate_reason[path] = "identical file hash"
    for stem, group in stem_groups.items():
        extensions = {path.suffix.lower() for path in group}
        if len(group) > 1 and len(extensions) > 1:
            for path in group:
                duplicate_group[path] = duplicate_group[path] or f"same-title:{slugify(stem, 60)}"
                duplicate_reason[path] = duplicate_reason[path] or "same title with different format"
    return hashes, duplicate_group, duplicate_reason


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse recently downloaded safety forms into Supabase dry-run payloads.")
    parser.add_argument("--downloads-dir", default=str(Path.home() / "Downloads"))
    parser.add_argument("--out-dir", default="evaluation/download-safety-bundle-2026-05-03")
    parser.add_argument("--since", default="2026-05-03T18:45:00", help="Local timestamp lower bound, or empty for all supported files.")
    parser.add_argument("--max-pdf-pages", type=int, default=8)
    parser.add_argument("--max-sheet-rows", type=int, default=60)
    args = parser.parse_args()

    started_at = time.perf_counter()
    downloads_dir = Path(args.downloads_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    since = datetime.fromisoformat(args.since) if args.since else None
    target_files = find_target_files(downloads_dir, since)
    hashes, duplicate_groups, duplicate_reasons = duplicate_maps(target_files)

    inventory: list[InventoryEntry] = []
    results: list[ParseResult] = []
    sources: list[SourcePayload] = []
    items: list[ItemPayload] = []

    for path in target_files:
        stat = path.stat()
        entry = InventoryEntry(
            id=f"file-{hashes[path][:12]}",
            path=str(path),
            name=path.name,
            extension=path.suffix.lower(),
            size_bytes=stat.st_size,
            modified_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
            sha256=hashes[path],
            duplicate_group=duplicate_groups[path],
            duplicate_reason=duplicate_reasons[path],
        )
        result = parse_file(path, args.max_pdf_pages, args.max_sheet_rows)
        source = build_source(path, entry, result)
        item = build_item(result)
        inventory.append(entry)
        results.append(result)
        sources.append(source)
        if item:
            items.append(item)

    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
    failures = [result for result in results if not result.ok]
    run_record = {
        "source_batch": "downloads-safety-forms-2026-05-03",
        "source_count": len(sources),
        "item_count": len(items),
        "success_count": len(results) - len(failures),
        "failure_count": len(failures),
        "elapsed_ms": elapsed_ms,
        "report_path": str(out_dir / "report.md"),
        "status": "completed_with_notice" if failures else "completed",
        "details": {
            "downloadsDir": str(downloads_dir),
            "since": args.since,
            "duplicateCount": sum(1 for entry in inventory if entry.duplicate_group),
            "failureNames": [result.name for result in failures],
            "noDbMutation": True,
        },
    }
    recommended_schema = {
        "safety_reference_sources": [
            "id",
            "source_group",
            "source_type",
            "agency",
            "title",
            "source_path",
            "origin_url",
            "file_format",
            "published_at",
            "metadata",
        ],
        "safety_reference_items": [
            "id",
            "source_id",
            "item_type",
            "category",
            "subcategory",
            "title",
            "summary",
            "body",
            "keywords",
            "risk_tags",
            "primary_documents",
            "controls",
            "payload",
        ],
        "safety_reference_ingestion_runs": [
            "source_batch",
            "source_count",
            "item_count",
            "success_count",
            "failure_count",
            "elapsed_ms",
            "report_path",
            "status",
            "details",
        ],
        "recommended_additions": [
            "For scanned PDF/image-heavy sources, keep OCR status in source.metadata.ocrStatus.",
            "For form templates, keep extracted section and signature fields in item.payload.formFields.",
            "For versioned uploads, keep sha256 in source.metadata and use it for duplicate checks before upsert.",
        ],
    }
    sample_items = items[:5]
    report = {
        "generated_at": now_iso(),
        "run": run_record,
        "inventory": [asdict(entry) for entry in inventory],
        "parse_results": [asdict(result) for result in results],
        "recommended_schema": recommended_schema,
        "duplicates": [
            asdict(entry)
            for entry in inventory
            if entry.duplicate_group
        ],
        "samples": [asdict(item) for item in sample_items],
        "upsert_payload": {
            "sources": [asdict(source) for source in sources],
            "items": [asdict(item) for item in items],
            "ingestion_run": run_record,
        },
    }

    write_json(out_dir / "inventory.json", [asdict(entry) for entry in inventory])
    write_json(out_dir / "parse-report.json", report)
    write_json(out_dir / "recommended-schema.json", recommended_schema)
    write_json(out_dir / "samples.json", [asdict(item) for item in sample_items])
    write_json(out_dir / "upsert-payload.json", report["upsert_payload"])
    (out_dir / "upsert.sql").write_text(build_upsert_sql(sources, items, run_record), encoding="utf-8")

    markdown_lines = [
        "# Downloads safety form parsing bundle",
        "",
        f"- generated_at: {report['generated_at']}",
        f"- downloads_dir: {downloads_dir}",
        f"- since: {args.since}",
        f"- source_count: {run_record['source_count']}",
        f"- success_count: {run_record['success_count']}",
        f"- failure_count: {run_record['failure_count']}",
        f"- item_count: {run_record['item_count']}",
        f"- duplicate_count: {run_record['details']['duplicateCount']}",
        f"- db_mutation: none",
        "",
        "## Inventory",
        "",
        "| name | ext | size | duplicate | parse | form_type |",
        "|---|---:|---:|---|---|---|",
    ]
    for entry, result in zip(inventory, results):
        duplicate = entry.duplicate_group or ""
        status = "ok" if result.ok else f"fail: {result.failure_reason}"
        markdown_lines.append(f"| {entry.name} | {entry.extension} | {entry.size_bytes} | {duplicate} | {status} | {result.form_type} |")
    markdown_lines.extend([
        "",
        "## Recommended Tables / Columns",
        "",
        "- `safety_reference_sources`: source metadata, file hash, duplicate group, parser status",
        "- `safety_reference_items`: searchable item body, keywords, risk tags, SafeClaw primary document mapping, controls",
        "- `safety_reference_ingestion_runs`: dry-run counts, elapsed time, report path, failure list",
        "",
        "## Duplicate Notes",
        "",
    ])
    if report["duplicates"]:
        markdown_lines.extend([f"- {entry['name']}: {entry['duplicate_reason']} ({entry['duplicate_group']})" for entry in report["duplicates"]])
    else:
        markdown_lines.append("- none")
    markdown_lines.extend([
        "",
        "## Sample Items",
        "",
    ])
    for item in sample_items:
        markdown_lines.extend([
            f"### {item.title}",
            f"- id: `{item.id}`",
            f"- item_type: {item.item_type}",
            f"- primary_documents: {', '.join(item.primary_documents)}",
            f"- keywords: {', '.join(item.keywords)}",
            f"- summary: {item.summary}",
            "",
        ])
    markdown_lines.extend([
        "## Remaining Gaps",
        "",
        "- `.hwp` and `.xls` binary originals are inventoried but not text-parsed in this dry-run; matching `.hwpx` files cover the uploaded TBM/form templates where available.",
        "- PDF parsing uses text extraction from the first configured pages; long manuals should be chunked page-by-page before production ingestion.",
        "- Image/OCR-only PDF quality is flagged through short text length and should be OCR-verified before DB load.",
        "",
        "## Generated Files",
        "",
        "- `inventory.json`",
        "- `parse-report.json`",
        "- `recommended-schema.json`",
        "- `samples.json`",
        "- `upsert-payload.json`",
        "- `upsert.sql`",
    ])
    (out_dir / "report.md").write_text("\n".join(markdown_lines), encoding="utf-8")
    print(json.dumps(run_record, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
