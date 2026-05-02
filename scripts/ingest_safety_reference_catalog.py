from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import time
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence
from urllib import error, request

from openpyxl import load_workbook
from pypdf import PdfReader


DOCUMENTS_RISK = ["점검결과 요약", "위험성평가표", "작업계획서"]
DOCUMENTS_TBM = ["TBM 브리핑", "TBM 기록", "안전보건교육 기록"]
DOCUMENTS_EVIDENCE = ["사진/증빙", "비상대응 절차"]
RISK_KEYWORDS = [
    "추락",
    "끼임",
    "충돌",
    "전도",
    "화재",
    "폭발",
    "질식",
    "중독",
    "감전",
    "화학",
    "비계",
    "굴착",
    "크레인",
    "지게차",
    "밀폐",
    "고소",
    "폭염",
    "자외선",
    "온열",
    "절단",
    "낙하",
    "붕괴",
]


@dataclass(frozen=True)
class ReferenceSource:
    id: str
    source_group: str
    source_type: str
    agency: str
    title: str
    source_path: str
    origin_url: str | None
    file_format: str
    published_at: str | None
    metadata: dict[str, object]


@dataclass(frozen=True)
class ReferenceItem:
    id: str
    source_id: str
    item_type: str
    category: str | None
    subcategory: str | None
    title: str
    summary: str
    body: str
    keywords: list[str]
    risk_tags: list[str]
    primary_documents: list[str]
    controls: list[str]
    payload: dict[str, object]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def slugify(value: str, max_length: int = 96) -> str:
    normalized = re.sub(r"[^0-9A-Za-z가-힣]+", "-", value.strip()).strip("-").lower()
    if not normalized:
        digest = hashlib.sha1(value.encode("utf-8", errors="ignore")).hexdigest()[:12]
        return f"item-{digest}"
    if len(normalized) <= max_length:
        return normalized
    digest = hashlib.sha1(value.encode("utf-8", errors="ignore")).hexdigest()[:10]
    return f"{normalized[:max_length - 11].rstrip('-')}-{digest}"


def compact_text(value: object, limit: int = 700) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit].rstrip()


def read_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def infer_keywords(*values: str) -> list[str]:
    text = " ".join(values)
    return [keyword for keyword in RISK_KEYWORDS if keyword in text]


def infer_documents(*values: str) -> list[str]:
    text = " ".join(values)
    documents: list[str] = []
    if re.search(r"공정|작업|설비|기계|장비|공사|시공", text):
        documents.extend(DOCUMENTS_RISK)
    if re.search(r"교육|JSA|TBM|작업 전|안전보건|교안", text, re.IGNORECASE):
        documents.extend(DOCUMENTS_TBM)
    if re.search(r"증빙|사진|사고|재해|비상|대응|응급", text):
        documents.extend(DOCUMENTS_EVIDENCE)
    if not documents:
        documents.extend(["위험성평가표", "TBM 브리핑", "안전보건교육 기록"])
    return sorted(set(documents), key=documents.index)


def infer_controls(*values: str) -> list[str]:
    text = " ".join(values)
    controls: list[str] = []
    if re.search(r"추락|비계|고소|지붕", text):
        controls.extend(["작업발판·난간·개구부 상태 확인", "안전대 체결 및 작업반경 출입통제"])
    if re.search(r"끼임|협착|기계|설비|회전", text):
        controls.extend(["가동부 방호덮개와 비상정지장치 확인", "정비 전 전원 차단 및 잠금표지"])
    if re.search(r"지게차|충돌|차량|운반", text):
        controls.extend(["보행자 동선과 장비 동선 분리", "신호수 배치 및 후진 경보 확인"])
    if re.search(r"화학|물질|용제|독성|MSDS|세척|세정", text, re.IGNORECASE):
        controls.extend(["MSDS 확인 및 적정 보호구 착용", "환기와 누출 대응 물품 비치"])
    if re.search(r"밀폐|질식|산소|중독", text):
        controls.extend(["작업 전 산소·유해가스 농도 측정", "감시인 배치 및 구조장비 확보"])
    if re.search(r"폭염|자외선|온열|고온|옥외|실외", text):
        controls.extend(["물·그늘·휴식 계획 수립", "온열질환 이상징후 상호 확인"])
    if not controls:
        controls.extend(["작업 전 유해·위험요인 확인", "관리감독자 확인 후 작업 시작"])
    return sorted(set(controls), key=controls.index)


def read_csv_dicts(path: Path) -> tuple[list[dict[str, str]], str]:
    encodings = ["utf-8-sig", "cp949", "euc-kr"]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                rows = list(csv.DictReader(file))
            return rows, encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"CSV 인코딩을 확인할 수 없습니다: {path} ({last_error})")


def first_non_empty_header(rows: Iterable[tuple[object, ...]]) -> tuple[list[str], list[tuple[object, ...]]]:
    consumed: list[tuple[object, ...]] = []
    header: list[str] = []
    for row in rows:
        values = [compact_text(cell, 120) for cell in row]
        if sum(1 for value in values if value) >= 3:
            header = [value or f"column_{index + 1}" for index, value in enumerate(values)]
            break
    for row in rows:
        consumed.append(row)
    return header, consumed


def parse_sif_archive(path: Path) -> tuple[ReferenceSource, list[ReferenceItem]]:
    source = ReferenceSource(
        id="kosha-sif-archive-20260401",
        source_group="kosha-reference",
        source_type="xlsx",
        agency="한국산업안전보건공단",
        title="산업재해 고위험요인(SIF) 아카이브",
        source_path=str(path),
        origin_url=None,
        file_format="xlsx",
        published_at="2026-04-01",
        metadata={"parser": "openpyxl", "fileName": path.name},
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    items: list[ReferenceItem] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)
        header, data_rows = first_non_empty_header(rows_iter)
        for index, row in enumerate(data_rows, start=1):
            record = {
                header[column_index]: compact_text(row[column_index] if column_index < len(row) else "", 500)
                for column_index in range(len(header))
            }
            if not any(record.values()):
                continue
            title_seed = " / ".join(value for value in list(record.values())[:4] if value) or f"{sheet_name} {index}"
            title = compact_text(title_seed, 180)
            body = "\n".join(f"{key}: {value}" for key, value in record.items() if value)
            risk_tags = infer_keywords(title, body)
            items.append(ReferenceItem(
                id=f"sif-{slugify(sheet_name)}-{index:05d}",
                source_id=source.id,
                item_type="sif-case",
                category=sheet_name,
                subcategory=None,
                title=title,
                summary=compact_text(body, 260),
                body=body,
                keywords=sorted(set(risk_tags + [sheet_name])),
                risk_tags=risk_tags,
                primary_documents=infer_documents(title, body),
                controls=infer_controls(title, body),
                payload={"sheet": sheet_name, "rowIndex": index, "record": record},
            ))
    return source, items


def parse_construction_process(path: Path) -> tuple[ReferenceSource, list[ReferenceItem]]:
    rows, encoding = read_csv_dicts(path)
    source = ReferenceSource(
        id="kosha-construction-process-20210910",
        source_group="kosha-reference",
        source_type="csv",
        agency="한국산업안전보건공단",
        title="건설업 공종별 세부공정 목록",
        source_path=str(path),
        origin_url=None,
        file_format="csv",
        published_at="2021-09-10",
        metadata={"encoding": encoding, "fileName": path.name},
    )
    items: list[ReferenceItem] = []
    for index, row in enumerate(rows, start=1):
        construction_type = compact_text(row.get("공사종류", ""), 120)
        work_type = compact_text(row.get("공종명", ""), 120)
        detail = compact_text(row.get("세부공정명", ""), 160)
        title = " · ".join(part for part in [construction_type, work_type, detail] if part) or f"건설 세부공정 {index}"
        body = f"공사종류: {construction_type}\n공종명: {work_type}\n세부공정명: {detail}"
        items.append(ReferenceItem(
            id=f"construction-process-{index:04d}",
            source_id=source.id,
            item_type="construction-process",
            category=construction_type or None,
            subcategory=work_type or None,
            title=title,
            summary=f"{construction_type} 공사에서 {work_type} / {detail} 공정을 작업계획서와 위험성평가 분류 기준으로 사용합니다.",
            body=body,
            keywords=[value for value in [construction_type, work_type, detail] if value],
            risk_tags=infer_keywords(title),
            primary_documents=["작업계획서", "위험성평가표", "TBM 브리핑"],
            controls=infer_controls(title),
            payload={"record": row},
        ))
    return source, items


def parse_machinery(path: Path) -> tuple[ReferenceSource, list[ReferenceItem]]:
    rows, encoding = read_csv_dicts(path)
    source = ReferenceSource(
        id="kosha-machinery-20210909",
        source_group="kosha-reference",
        source_type="csv",
        agency="한국산업안전보건공단",
        title="업종별 기계설비 목록",
        source_path=str(path),
        origin_url=None,
        file_format="csv",
        published_at="2021-09-09",
        metadata={"encoding": encoding, "fileName": path.name},
    )
    items: list[ReferenceItem] = []
    for index, row in enumerate(rows, start=1):
        values = [compact_text(value, 140) for value in row.values() if compact_text(value, 140)]
        title = " · ".join(values[:4]) or f"업종별 기계설비 {index}"
        body = "\n".join(f"{key}: {value}" for key, value in row.items() if compact_text(value, 300))
        items.append(ReferenceItem(
            id=f"machinery-{index:04d}",
            source_id=source.id,
            item_type="machinery",
            category=values[0] if values else None,
            subcategory=values[1] if len(values) > 1 else None,
            title=title,
            summary=compact_text(body, 260),
            body=body,
            keywords=values[:6],
            risk_tags=infer_keywords(title, body),
            primary_documents=["작업계획서", "위험성평가표", "TBM 브리핑", "안전보건교육 기록"],
            controls=infer_controls(title, body),
            payload={"record": row},
        ))
    return source, items


def extract_pdf_text_from_bytes(data: bytes, max_pages: int) -> tuple[str, int]:
    reader = PdfReader(io.BytesIO(data))
    texts: list[str] = []
    page_count = len(reader.pages)
    for page in reader.pages[:max_pages]:
        texts.append(page.extract_text() or "")
    return "\n".join(texts).strip(), page_count


def extract_pdf_text(path: Path, max_pages: int) -> tuple[str, int]:
    with path.open("rb") as file:
        return extract_pdf_text_from_bytes(file.read(), max_pages)


def parse_pdf_source(path: Path, source_id: str, title: str, item_type: str, max_pages: int) -> tuple[ReferenceSource, list[ReferenceItem]]:
    text, page_count = extract_pdf_text(path, max_pages)
    source = ReferenceSource(
        id=source_id,
        source_group="kosha-reference",
        source_type="pdf",
        agency="한국산업안전보건공단",
        title=title,
        source_path=str(path),
        origin_url=None,
        file_format="pdf",
        published_at=None,
        metadata={"fileName": path.name, "pageCount": page_count, "parsedPages": min(max_pages, page_count)},
    )
    summary = compact_text(text, 420) or "PDF 원문 이미지 비중이 높아 OCR 보강이 필요한 자료입니다."
    item = ReferenceItem(
        id=f"{source_id}-summary",
        source_id=source.id,
        item_type=item_type,
        category="위험성평가",
        subcategory=None,
        title=title,
        summary=summary,
        body=text[:6000],
        keywords=infer_keywords(title, text) + ["위험성평가"],
        risk_tags=infer_keywords(title, text),
        primary_documents=infer_documents(title, text),
        controls=infer_controls(title, text),
        payload={"pageCount": page_count, "parsedPages": min(max_pages, page_count), "needsOcr": len(text) < 500},
    )
    return source, [item]


def decode_zip_name(name: str) -> str:
    try:
        repaired = name.encode("cp437").decode("cp949")
        if sum(1 for char in repaired if "\uac00" <= char <= "\ud7a3") >= sum(1 for char in name if "\uac00" <= char <= "\ud7a3"):
            return repaired
    except UnicodeError:
        pass
    return name


def parse_technical_support_zips(folder: Path, max_pdf_pages: int, priority_only: bool) -> tuple[ReferenceSource, list[ReferenceItem]]:
    source = ReferenceSource(
        id="kosha-technical-support-regulations-2025",
        source_group="kosha-reference",
        source_type="zip-folder",
        agency="한국산업안전보건공단",
        title="기술지원규정 및 안전보건 기술지침 묶음",
        source_path=str(folder),
        origin_url=None,
        file_format="zip/pdf",
        published_at="2025-01-01",
        metadata={"folder": str(folder), "priorityOnly": priority_only},
    )
    items: list[ReferenceItem] = []
    zip_paths = sorted(folder.glob("*.zip"))
    for zip_index, zip_path in enumerate(zip_paths, start=1):
        with zipfile.ZipFile(zip_path) as archive:
            for file_index, info in enumerate(archive.infolist(), start=1):
                if info.is_dir():
                    continue
                decoded_name = decode_zip_name(info.filename)
                if not decoded_name.lower().endswith(".pdf"):
                    continue
                is_priority = "기술지원규정" in decoded_name
                if priority_only and not is_priority:
                    continue
                text = ""
                page_count = 0
                needs_ocr = False
                if is_priority:
                    try:
                        data = archive.read(info)
                        text, page_count = extract_pdf_text_from_bytes(data, max_pdf_pages)
                        needs_ocr = len(text) < 500
                    except Exception as exc:
                        text = ""
                        needs_ocr = True
                        page_count = 0
                        print(f"[warn] PDF text extraction failed: {decoded_name} ({exc})")
                title = Path(decoded_name).stem
                category = zip_path.stem.replace("[2025] 기술지원규정(", "").replace(").zip", "").replace(")", "")
                summary = compact_text(text, 420) if text else f"{category} 분야의 KOSHA 기술지원규정 또는 안전보건 기술지침 자료입니다."
                items.append(ReferenceItem(
                    id=f"technical-support-{zip_index:02d}-{file_index:04d}-{slugify(title, 60)}",
                    source_id=source.id,
                    item_type="technical-support-regulation" if is_priority else "technical-guideline",
                    category=category,
                    subcategory="기술지원규정" if is_priority else "기술지침",
                    title=title,
                    summary=summary,
                    body=text[:6000],
                    keywords=sorted(set(infer_keywords(title, text) + [category, "KOSHA"])),
                    risk_tags=infer_keywords(title, text),
                    primary_documents=infer_documents(title, text),
                    controls=infer_controls(title, text),
                    payload={
                        "zipFile": zip_path.name,
                        "internalPath": decoded_name,
                        "compressedSize": info.compress_size,
                        "fileSize": info.file_size,
                        "isPriority": is_priority,
                        "pageCount": page_count,
                        "parsedPages": min(max_pdf_pages, page_count) if page_count else 0,
                        "needsOcr": needs_ocr,
                    },
                ))
    return source, items


def chunked(items: Sequence[dict[str, object]], size: int) -> Iterable[list[dict[str, object]]]:
    for index in range(0, len(items), size):
        yield list(items[index:index + size])


def supabase_upsert(table: str, rows: Sequence[dict[str, object]], chunk_size: int = 500) -> int:
    if not rows:
        return 0
    supabase_url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required for upload")
    total = 0
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/{table}?on_conflict=id"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    for rows_chunk in chunked(rows, chunk_size):
        data = json.dumps(rows_chunk, ensure_ascii=False).encode("utf-8")
        req = request.Request(endpoint, data=data, headers=headers, method="POST")
        try:
            with request.urlopen(req, timeout=60) as response:
                if response.status not in (200, 201, 204):
                    raise RuntimeError(f"Unexpected status {response.status} for {table}")
        except error.HTTPError as exc:
            message = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase upsert failed for {table}: {exc.code} {message}") from exc
        total += len(rows_chunk)
    return total


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build and optionally upload SafeClaw safety reference catalog.")
    parser.add_argument("--downloads-dir", default=r"C:\Users\iceam\Downloads")
    parser.add_argument("--technical-folder", default=r"C:\Users\iceam\Downloads\기술지원규정")
    parser.add_argument("--output-dir", default="evaluation/data-ingestion")
    parser.add_argument("--max-pdf-pages", type=int, default=3)
    parser.add_argument("--include-secondary-technical", action="store_true", help="Include non-priority guideline PDFs from technical ZIPs.")
    parser.add_argument("--upload", action="store_true", help="Upload parsed records to Supabase REST after migration is applied.")
    args = parser.parse_args()

    started = time.perf_counter()
    read_env_file(Path(".env.local"))
    downloads = Path(args.downloads_dir)
    output_dir = Path(args.output_dir)
    sources: list[ReferenceSource] = []
    items: list[ReferenceItem] = []
    failures: list[dict[str, str]] = []

    jobs = [
        ("sif", lambda: parse_sif_archive(downloads / "한국산업안전보건공단_산업재해 고위험요인(SIF) 아카이브_20260401.xlsx")),
        ("construction", lambda: parse_construction_process(downloads / "한국산업안전보건공단_건설업 공종별 세부공정 목록_20210910.csv")),
        ("machinery", lambda: parse_machinery(downloads / "한국산업안전보건공단_업종별 기계설비 목록_20210909.csv")),
        ("risk-manual", lambda: parse_pdf_source(downloads / "위험성평가 접근방법 메뉴얼.pdf", "kosha-risk-assessment-approach-manual", "위험성평가 접근방법 메뉴얼", "risk-manual", args.max_pdf_pages)),
        ("jsa-training", lambda: parse_pdf_source(downloads / "위험성평가 JSA 교육 교안.pdf", "kosha-jsa-training-deck", "위험성평가 JSA 교육 교안", "jsa-training", args.max_pdf_pages)),
        ("technical-support", lambda: parse_technical_support_zips(Path(args.technical_folder), args.max_pdf_pages, not args.include_secondary_technical)),
    ]

    for label, job in jobs:
        try:
            source, parsed_items = job()
            sources.append(source)
            items.extend(parsed_items)
            print(f"[ok] {label}: sources=1 items={len(parsed_items)}")
        except Exception as exc:
            failures.append({"job": label, "message": str(exc)})
            print(f"[fail] {label}: {exc}")

    source_rows = [asdict(source) for source in sources]
    item_rows = [asdict(item) for item in items]
    seed = {
        "generatedAt": now_iso(),
        "sources": source_rows,
        "items": item_rows,
    }
    seed_path = output_dir / "safety-reference-catalog.seed.json"
    write_json(seed_path, seed)

    uploaded = {"sources": 0, "items": 0, "runs": 0}
    upload_error: str | None = None
    if args.upload:
        try:
            uploaded["sources"] = supabase_upsert("safety_reference_sources", source_rows)
            uploaded["items"] = supabase_upsert("safety_reference_items", item_rows)
        except Exception as exc:
            upload_error = str(exc)
            print(f"[warn] upload failed: {upload_error}")

    elapsed_ms = round((time.perf_counter() - started) * 1000)
    report = {
        "generatedAt": now_iso(),
        "elapsedMs": elapsed_ms,
        "sourceCount": len(sources),
        "itemCount": len(items),
        "successCount": len(jobs) - len(failures),
        "failureCount": len(failures),
        "failures": failures,
        "uploaded": uploaded,
        "uploadError": upload_error,
        "seedPath": str(seed_path),
        "itemTypeCounts": {},
        "documentCounts": {},
        "riskTagCounts": {},
    }
    for item in items:
        report["itemTypeCounts"][item.item_type] = report["itemTypeCounts"].get(item.item_type, 0) + 1
        for document in item.primary_documents:
            report["documentCounts"][document] = report["documentCounts"].get(document, 0) + 1
        for tag in item.risk_tags:
            report["riskTagCounts"][tag] = report["riskTagCounts"].get(tag, 0) + 1

    report_path = output_dir / "safety-reference-catalog-report.json"
    write_json(report_path, report)

    if args.upload and upload_error is None:
        run_row = {
            "source_batch": "safety-reference-catalog-2026-05-02",
            "source_count": len(sources),
            "item_count": len(items),
            "success_count": report["successCount"],
            "failure_count": report["failureCount"],
            "elapsed_ms": elapsed_ms,
            "report_path": str(report_path),
            "status": "completed",
            "details": report,
        }
        try:
            uploaded["runs"] = supabase_upsert("safety_reference_ingestion_runs", [run_row])
            report["uploaded"] = uploaded
            write_json(report_path, report)
        except Exception as exc:
            report["uploadError"] = str(exc)
            write_json(report_path, report)
            print(f"[warn] run upload failed: {exc}")

    print(f"[done] sources={len(sources)} items={len(items)} failures={len(failures)} elapsedMs={elapsed_ms}")
    print(f"[artifact] {report_path}")
    print(f"[artifact] {seed_path}")
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
