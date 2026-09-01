from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from PIL import Image
SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from archive_safety import (
    ArchiveBudgetError,
    ArchiveLimits,
    BoundedZipReader,
    preflight_zip_central_directory,
)
from pdf_parser_worker import parse_pdf_file_bounded


DEFAULT_MAX_FILES = 10_000
DEFAULT_MAX_TOTAL_BYTES = 4 * 1024 * 1024 * 1024
DEFAULT_MAX_FILE_BYTES = 128 * 1024 * 1024
DEFAULT_MAX_PARSER_FILES = 2_000
DEFAULT_MAX_ELAPSED_SECONDS = 900.0
DEFAULT_MAX_IMAGE_PIXELS = 80_000_000
STRUCTURED_ARCHIVE_LIMITS = ArchiveLimits(
    max_member_count=4_096,
    max_member_bytes=64 * 1024 * 1024,
    max_total_uncompressed_bytes=512 * 1024 * 1024,
    max_compression_ratio=100.0,
    max_central_directory_bytes=16 * 1024 * 1024,
)


KEYWORDS: dict[str, list[str]] = {
    "risk_assessment": ["위험성평가", "유해", "위험요인", "빈도", "강도", "감소대책", "개선대책"],
    "tbm_prework": ["TBM", "작업 전", "작업전", "안전점검회의", "일일안전", "공사감독일지", "안전감독일지"],
    "work_plan": ["작업계획서", "표준 작업계획서", "작업순서", "작업방법", "열수송관", "굴착"],
    "emergency_response": ["중대재해", "비상", "대응", "사고발생", "응급", "보고체계"],
    "education": ["교육", "협력회사", "전달", "근로자", "안전보건교육"],
    "contractor_owner": ["발주자", "시공자", "도급", "수급", "협력사"],
    "inspection": ["점검", "체크", "확인", "평가표", "수준 평가"],
    "photo_evidence": ["사진", "강평", "개선", "작업중지권", "신호수"],
}


@dataclass
class FileRecord:
    path: str
    relativePath: str
    extension: str
    size: int
    modified: str
    categoryHints: list[str]


@dataclass(frozen=True)
class ScanLimits:
    max_files: int = DEFAULT_MAX_FILES
    max_total_bytes: int = DEFAULT_MAX_TOTAL_BYTES
    max_file_bytes: int = DEFAULT_MAX_FILE_BYTES
    max_parser_files: int = DEFAULT_MAX_PARSER_FILES
    max_elapsed_seconds: float = DEFAULT_MAX_ELAPSED_SECONDS
    max_image_pixels: int = DEFAULT_MAX_IMAGE_PIXELS


@dataclass(frozen=True)
class FileCandidate:
    path: Path
    size: int
    modified_ns: int


class ScanBudgetError(RuntimeError):
    pass


def _assert_before_deadline(started: float, limits: ScanLimits, stage: str) -> None:
    elapsed = time.perf_counter() - started
    if elapsed > limits.max_elapsed_seconds:
        raise ScanBudgetError(
            f"scan elapsed time exceeds limit at {stage}: "
            f"{elapsed:.3f}/{limits.max_elapsed_seconds:.3f} seconds"
        )


def discover_files(source: Path, limits: ScanLimits, started: float) -> tuple[list[FileCandidate], int]:
    if not source.exists():
        raise ScanBudgetError(f"source directory does not exist: {source}")
    if not source.is_dir():
        raise ScanBudgetError(f"source path is not a directory: {source}")
    if source.is_symlink():
        raise ScanBudgetError(f"source directory must not be a symlink: {source}")

    files: list[FileCandidate] = []
    directories = [source]
    total_bytes = 0
    while directories:
        _assert_before_deadline(started, limits, "directory traversal")
        current = directories.pop()
        try:
            with os.scandir(current) as entries:
                for entry in entries:
                    _assert_before_deadline(started, limits, "directory traversal")
                    if entry.is_symlink():
                        continue
                    if entry.is_dir(follow_symlinks=False):
                        directories.append(Path(entry.path))
                        continue
                    if not entry.is_file(follow_symlinks=False):
                        continue
                    stat = entry.stat(follow_symlinks=False)
                    if stat.st_size > limits.max_file_bytes:
                        raise ScanBudgetError(
                            f"file size exceeds limit: {entry.path} "
                            f"({stat.st_size}/{limits.max_file_bytes})"
                        )
                    if len(files) >= limits.max_files:
                        raise ScanBudgetError(
                            f"file count exceeds limit: {len(files) + 1}/{limits.max_files}"
                        )
                    total_bytes += stat.st_size
                    if total_bytes > limits.max_total_bytes:
                        raise ScanBudgetError(
                            f"aggregate source bytes exceed limit: "
                            f"{total_bytes}/{limits.max_total_bytes}"
                        )
                    files.append(FileCandidate(Path(entry.path), stat.st_size, stat.st_mtime_ns))
        except OSError as exc:
            raise ScanBudgetError(f"failed to traverse source directory {current}: {exc}") from exc
    files.sort(key=lambda item: item.path.as_posix().lower())
    return files, total_bytes


def preflight_structured_document(path: Path) -> None:
    try:
        preflight_zip_central_directory(path, STRUCTURED_ARCHIVE_LIMITS)
        with zipfile.ZipFile(path) as archive:
            BoundedZipReader(archive, STRUCTURED_ARCHIVE_LIMITS)
    except (ArchiveBudgetError, OSError) as exc:
        raise ScanBudgetError(f"structured document archive preflight failed for {path}: {exc}") from exc


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def category_hints(text: str) -> list[str]:
    hits: list[str] = []
    for category, words in KEYWORDS.items():
        if any(word.lower() in text.lower() for word in words):
            hits.append(category)
    return hits


def short_text(value: str, limit: int = 120) -> str:
    normalized = normalize_text(value)
    if len(normalized) <= limit:
        return normalized
    return f"{normalized[:limit].rstrip()}..."


def read_zip_xml_text(path: Path, xml_suffixes: tuple[str, ...]) -> list[str]:
    texts: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            bounded_archive = BoundedZipReader(archive)
            for info in bounded_archive.infos:
                name = info.filename
                if not name.lower().endswith(xml_suffixes):
                    continue
                try:
                    root = ElementTree.fromstring(bounded_archive.read(info))
                except ElementTree.ParseError:
                    continue
                for node in root.iter():
                    if node.text and normalize_text(node.text):
                        texts.append(normalize_text(node.text))
    except (OSError, zipfile.BadZipFile):
        return []
    return texts


def inspect_xlsx(path: Path) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - report scanner errors without failing the full run.
        return {"path": str(path), "error": str(exc)}

    sheets: list[dict[str, Any]] = []
    total_non_empty = 0
    for sheet in workbook.worksheets[:12]:
        rows: list[list[str]] = []
        non_empty = 0
        for row in sheet.iter_rows(max_row=25, max_col=12, values_only=True):
            values = [short_text(str(value)) if value is not None else "" for value in row]
            if any(values):
                non_empty += 1
                rows.append(values)
            if len(rows) >= 8:
                break
        total_non_empty += non_empty
        joined = " ".join(" ".join(row) for row in rows)
        sheets.append(
            {
                "name": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "sampleRows": rows,
                "categoryHints": category_hints(f"{sheet.title} {joined}"),
            }
        )
    workbook.close()
    return {
        "path": str(path),
        "sheetCount": len(workbook.sheetnames),
        "sheetNames": workbook.sheetnames[:30],
        "sampledSheets": sheets,
        "sampledNonEmptyRows": total_non_empty,
        "categoryHints": category_hints(f"{path.name} {' '.join(workbook.sheetnames)}"),
    }


def inspect_hwpx(path: Path) -> dict[str, Any]:
    texts = read_zip_xml_text(path, (".xml",))
    joined = " ".join(texts)
    headings = [short_text(text) for text in texts if 3 <= len(text) <= 80]
    return {
        "path": str(path),
        "textNodeCount": len(texts),
        "sampleHeadings": headings[:30],
        "categoryHints": category_hints(f"{path.name} {joined[:3000]}"),
    }


def inspect_pptx(path: Path) -> dict[str, Any]:
    texts = read_zip_xml_text(path, (".xml",))
    joined = " ".join(texts)
    return {
        "path": str(path),
        "textNodeCount": len(texts),
        "sampleTexts": [short_text(text) for text in texts[:60]],
        "categoryHints": category_hints(f"{path.name} {joined[:3000]}"),
    }


def inspect_pdf(path: Path, limits: ScanLimits, started: float) -> dict[str, Any]:
    try:
        remaining_seconds = limits.max_elapsed_seconds - (time.perf_counter() - started)
        if remaining_seconds <= 0:
            raise ScanBudgetError("scan elapsed time exceeds limit before PDF parser")
        parsed = parse_pdf_file_bounded(
            path,
            max_input_bytes=limits.max_file_bytes,
            extract_pages=5,
            max_total_pages=2_000,
            max_text_chars=2_000_000,
            timeout_seconds=min(30.0, remaining_seconds),
        )
        page_count = parsed.page_count
        snippets: list[str] = []
        for page in parsed.pages:
            text = page.text
            for line in text.splitlines():
                clean = normalize_text(line)
                if len(clean) >= 3:
                    snippets.append(short_text(clean))
                if len(snippets) >= 30:
                    break
            if len(snippets) >= 30:
                break
    except Exception as exc:  # noqa: BLE001
        return {"path": str(path), "error": str(exc)}
    return {
        "path": str(path),
        "pageCount": page_count,
        "sampleTexts": snippets,
        "categoryHints": category_hints(f"{path.name} {' '.join(snippets)}"),
    }


def inspect_image(path: Path, max_image_pixels: int = DEFAULT_MAX_IMAGE_PIXELS) -> dict[str, Any]:
    try:
        with Image.open(path) as image:
            width, height = image.size
            if width * height > max_image_pixels:
                raise ScanBudgetError(
                    f"image pixels exceed limit: {path} "
                    f"({width * height}/{max_image_pixels})"
                )
            mode = image.mode
    except ScanBudgetError:
        raise
    except Exception as exc:  # noqa: BLE001
        return {"path": str(path), "error": str(exc)}
    return {
        "path": str(path),
        "width": width,
        "height": height,
        "mode": mode,
        "megapixels": round((width * height) / 1_000_000, 2),
        "categoryHints": category_hints(path.name),
    }


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan industrial safety template sources.")
    parser.add_argument("--source", required=True, help="Source directory to scan.")
    parser.add_argument("--out", required=True, help="Output directory for scan artifacts.")
    parser.add_argument("--max-files", type=int, default=DEFAULT_MAX_FILES)
    parser.add_argument("--max-total-bytes", type=int, default=DEFAULT_MAX_TOTAL_BYTES)
    parser.add_argument("--max-file-bytes", type=int, default=DEFAULT_MAX_FILE_BYTES)
    parser.add_argument("--max-parser-files", type=int, default=DEFAULT_MAX_PARSER_FILES)
    parser.add_argument("--max-elapsed-seconds", type=float, default=DEFAULT_MAX_ELAPSED_SECONDS)
    parser.add_argument("--max-image-pixels", type=int, default=DEFAULT_MAX_IMAGE_PIXELS)
    args = parser.parse_args()

    started = time.perf_counter()
    source_argument = Path(args.source)
    if source_argument.is_symlink():
        raise ScanBudgetError(f"source directory must not be a symlink: {source_argument}")
    source = source_argument.resolve()
    out = Path(args.out)
    limits = ScanLimits(
        max_files=args.max_files,
        max_total_bytes=args.max_total_bytes,
        max_file_bytes=args.max_file_bytes,
        max_parser_files=args.max_parser_files,
        max_elapsed_seconds=args.max_elapsed_seconds,
        max_image_pixels=args.max_image_pixels,
    )
    if any(
        value <= 0
        for value in (
            limits.max_files,
            limits.max_total_bytes,
            limits.max_file_bytes,
            limits.max_parser_files,
            limits.max_elapsed_seconds,
            limits.max_image_pixels,
        )
    ):
        parser.error("all scan limits must be greater than zero")
    files, total_source_bytes = discover_files(source, limits, started)
    records: list[FileRecord] = []
    by_extension: Counter[str] = Counter()
    by_top_folder: Counter[str] = Counter()
    category_counts: Counter[str] = Counter()
    largest: list[dict[str, Any]] = []

    for candidate in files:
        _assert_before_deadline(started, limits, "inventory")
        path = candidate.path
        extension = path.suffix.lower() or "[none]"
        relative = path.relative_to(source)
        hint_text = f"{relative.as_posix()} {path.name}"
        hints = category_hints(hint_text)
        for hint in hints:
            category_counts[hint] += 1
        by_extension[extension] += 1
        parts = relative.parts
        by_top_folder[parts[0] if len(parts) > 1 else "[root]"] += 1
        records.append(
            FileRecord(
                path=str(path),
                relativePath=str(relative),
                extension=extension,
                size=candidate.size,
                modified=str(candidate.modified_ns),
                categoryHints=hints,
            )
        )
        largest.append({"path": str(path), "relativePath": str(relative), "size": candidate.size})

    largest = sorted(largest, key=lambda item: int(item["size"]), reverse=True)[:50]

    structured_docs: dict[str, list[dict[str, Any]]] = defaultdict(list)
    parser_file_count = 0
    for candidate in files:
        _assert_before_deadline(started, limits, "structured parser admission")
        path = candidate.path
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".hwpx", ".pdf", ".pptx"}:
            continue
        parser_file_count += 1
        if parser_file_count > limits.max_parser_files:
            raise ScanBudgetError(
                f"parser file count exceeds limit: "
                f"{parser_file_count}/{limits.max_parser_files}"
            )
        if suffix in {".xlsx", ".hwpx", ".pptx"}:
            preflight_structured_document(path)
        if suffix == ".xlsx":
            structured_docs["xlsx"].append(inspect_xlsx(path))
        elif suffix == ".hwpx":
            structured_docs["hwpx"].append(inspect_hwpx(path))
        elif suffix == ".pdf":
            structured_docs["pdf"].append(inspect_pdf(path, limits, started))
        elif suffix == ".pptx":
            structured_docs["pptx"].append(inspect_pptx(path))
        _assert_before_deadline(started, limits, f"parser completion for {path.name}")

    image_records: list[dict[str, Any]] = []
    for candidate in files:
        path = candidate.path
        if path.suffix.lower() not in {".jpg", ".jpeg", ".png", ".bmp"}:
            continue
        _assert_before_deadline(started, limits, "image parser admission")
        parser_file_count += 1
        if parser_file_count > limits.max_parser_files:
            raise ScanBudgetError(
                f"parser file count exceeds limit: "
                f"{parser_file_count}/{limits.max_parser_files}"
            )
        image_records.append(inspect_image(path, limits.max_image_pixels))
        _assert_before_deadline(started, limits, f"image parser completion for {path.name}")
    image_by_folder: Counter[str] = Counter()
    for item in image_records:
        raw_path = item.get("path")
        if not isinstance(raw_path, str):
            continue
        relative = Path(raw_path).relative_to(source)
        image_by_folder[relative.parts[0] if len(relative.parts) > 1 else "[root]"] += 1

    summary = {
        "sourceDirectory": str(source),
        "elapsedSeconds": round(time.perf_counter() - started, 2),
        "fileCount": len(files),
        "totalSourceBytes": total_source_bytes,
        "parserFileCount": parser_file_count,
        "limits": asdict(limits),
        "symlinkPolicy": "NO_FOLLOW",
        "byExtension": dict(by_extension.most_common()),
        "byTopFolder": dict(by_top_folder.most_common()),
        "categoryCounts": dict(category_counts.most_common()),
        "largestFiles": largest,
        "structuredCounts": {key: len(value) for key, value in structured_docs.items()},
        "imageCount": len(image_records),
        "imageByTopFolder": dict(image_by_folder.most_common()),
    }

    candidate_records = [
        asdict(record)
        for record in records
        if record.extension in {".xlsx", ".hwpx", ".pdf", ".pptx", ".hwp"}
        or record.categoryHints
    ]

    write_json(out / "summary.json", summary)
    write_json(out / "inventory.json", [asdict(record) for record in records])
    write_json(out / "candidate-records.json", candidate_records)
    write_json(out / "structured-documents.json", structured_docs)
    write_json(out / "image-inventory-summary.json", {"count": len(image_records), "samples": image_records[:120]})
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ScanBudgetError as exc:
        print(f"scan-budget-error: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc
